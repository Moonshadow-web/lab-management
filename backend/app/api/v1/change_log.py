from datetime import datetime
from io import BytesIO
import urllib.parse

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session

from ...core.crud_base import paginate
from ...core.database import get_db
from ...core.security import get_current_user
from ...models.file_change_log import FileChangeLog
from ...models.user import User

router = APIRouter(prefix="/change-logs", tags=["change-logs"])


class HandleUpdate(BaseModel):
    handled: bool


@router.get("")
def list_change_logs(
    page: int = 1,
    page_size: int = 20,
    q: str | None = None,
    change_type: str | None = None,
    handled: bool | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    query = db.query(FileChangeLog)
    if change_type:
        query = query.filter(FileChangeLog.change_type == change_type)
    if handled is not None:
        query = query.filter(FileChangeLog.handled == handled)
    if q:
        query = query.filter(
            or_(
                FileChangeLog.file_name.ilike(f"%{q}%"),
                FileChangeLog.file_code.ilike(f"%{q}%"),
                FileChangeLog.operator.ilike(f"%{q}%"),
            )
        )
    query = query.order_by(FileChangeLog.change_date.desc(), FileChangeLog.id.desc())
    return paginate(query, page, page_size)


@router.patch("/{log_id}")
def set_handled(log_id: int, body: HandleUpdate, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    """标记 / 撤销某条更改日志为已处理。"""
    log = db.get(FileChangeLog, log_id)
    if not log:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="未找到该日志记录")
    log.handled = body.handled
    log.handled_at = datetime.utcnow() if body.handled else None
    db.commit()
    db.refresh(log)
    return {
        "id": log.id,
        "handled": log.handled,
        "handled_at": log.handled_at.isoformat() if log.handled_at else None,
    }


@router.get("/export")
def export_change_logs(
    q: str | None = None,
    change_type: str | None = None,
    handled: bool | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导出「文件更改申请单」表格（xlsx）。"""
    query = db.query(FileChangeLog)
    if change_type:
        query = query.filter(FileChangeLog.change_type == change_type)
    if handled is not None:
        query = query.filter(FileChangeLog.handled == handled)
    if q:
        query = query.filter(
            or_(
                FileChangeLog.file_name.ilike(f"%{q}%"),
                FileChangeLog.file_code.ilike(f"%{q}%"),
                FileChangeLog.operator.ilike(f"%{q}%"),
            )
        )
    rows = query.order_by(FileChangeLog.change_date.desc(), FileChangeLog.id.desc()).all()
    return _build_xlsx(rows)


def _build_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "文件更改申请单"

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="4472C4")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=16, color="1F2329")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 标题
    ws.merge_cells("A1:G1")
    ws["A1"] = "文件更改申请单"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 导出日期
    ws.merge_cells("A2:G2")
    ws["A2"] = f"导出日期：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=10, color="909399")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")

    # 表头
    headers = ["序号", "文件名称", "文件编码", "更改类型", "更改日期", "申请人", "处理状态"]
    widths = [8, 40, 22, 12, 16, 16, 12]
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[3].height = 22

    # 数据
    for i, r in enumerate(rows, start=1):
        row_idx = 3 + i
        status_text = "已处理" if r.handled else "未处理"
        values = [i, r.file_name or "—", r.file_code or "—", r.change_type or "—", str(r.change_date), r.operator or "—", status_text]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=c, value=v)
            cell.border = border
            cell.alignment = left if c == 2 else center
            if c == 4:  # 更改类型着色
                color = {"新增": "C6EFCE", "修改": "FFEB9C", "作废": "FFC7CE"}.get(r.change_type)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
            if c == 7:  # 处理状态着色
                cell.fill = PatternFill("solid", fgColor="C6EFCE" if r.handled else "FCE4D6")
        ws.row_dimensions[row_idx].height = 18

    if not rows:
        ws.cell(row=4, column=1, value="（暂无更改记录）")
        ws.merge_cells("A4:G4")
        ws["A4"].alignment = center

    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"文件更改申请单_{datetime.now().strftime('%Y%m%d')}.xlsx"
    # filename= 仅放 ASCII（避免 latin-1 编码报错），中文名用 filename* 传递
    disp = f"attachment; filename=\"change_log.xlsx\"; filename*=UTF-8''{urllib.parse.quote(fname)}"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disp},
    )
