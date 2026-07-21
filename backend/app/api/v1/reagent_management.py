"""试剂管理：试剂目录/库存/盘库/订购/接收/月消耗 API"""

from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import or_, func
from sqlalchemy.orm import Session

from ...core.crud_base import paginate
from ...core.database import get_db
from ...core.security import get_current_user, require_roles
from ...models.reagent_management import (
    ReagentItem, ReagentStock, InventoryCheck, InventoryCheckItem,
    ReagentOrder, ReagentOrderItem, Receiving, ReceivingItem, ReagentConsumption,
    TestItemReagent, InstrumentReagent,
    detect_reagent_type, derive_library,
)
from ...models.test_item import TestItem
from ...models.instrument import Instrument
from ...models.user import User
from ...schemas.reagent_management import (
    ReagentItemCreate, ReagentItemUpdate, ReagentItemRead,
    ReagentStockRead, InventoryCheckCreate, InventoryCheckRead,
    ReagentOrderCreate, ReagentOrderRead, ReceivingCreate, ReceivingRead,
    ReagentConsumptionRead, ImportResult,
    TestItemReagentCreate, TestItemReagentRead,
    InstrumentReagentCreate, InstrumentReagentRead,
)

router = APIRouter(prefix="/reagent", tags=["reagent-management"])


# =============================================================================
# 1. 试剂目录 CRUD
# =============================================================================

@router.get("/items", response_model=dict)
def list_reagent_items(
    q: str = Query("", description="搜索（名称/品牌/材料编码）"),
    type: Optional[str] = Query(None, description="类型筛选"),
    category: Optional[str] = Query(None, description="类别筛选"),
    library: Optional[str] = Query(None, description="责任库筛选（生化凝血/免疫）"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    base = db.query(ReagentItem)
    if q.strip():
        kw = f"%{q.strip()}%"
        base = base.filter(
            or_(ReagentItem.name.like(kw), ReagentItem.brand.like(kw),
                ReagentItem.material_code.like(kw), ReagentItem.spec.like(kw))
        )
    if type:
        base = base.filter(ReagentItem.type == type)
    if category:
        base = base.filter(ReagentItem.category == category)
    if library:
        base = base.filter(ReagentItem.library == library)
    total = base.count()
    rows = base.order_by(ReagentItem.type, ReagentItem.category, ReagentItem.id).offset(
        (page - 1) * page_size
    ).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": [ReagentItemRead.model_validate(r) for r in rows]}


@router.get("/items/{item_id}", response_model=ReagentItemRead)
def get_reagent_item(item_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    item = db.query(ReagentItem).get(item_id)
    if not item:
        raise HTTPException(404, "试剂未找到")
    return item


@router.post("/items", response_model=ReagentItemRead)
def create_reagent_item(
    data: ReagentItemCreate, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    item = ReagentItem(**data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/items/{item_id}", response_model=ReagentItemRead)
def update_reagent_item(
    item_id: int, data: ReagentItemUpdate, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    item = db.query(ReagentItem).get(item_id)
    if not item:
        raise HTTPException(404, "试剂未找到")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/items/{item_id}")
def delete_reagent_item(
    item_id: int, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin")),
):
    item = db.query(ReagentItem).get(item_id)
    if not item:
        raise HTTPException(404, "试剂未找到")
    db.delete(item)
    db.commit()
    return {"ok": True}


# =============================================================================
# 2. 实时库存
# =============================================================================

@router.get("/stock", response_model=dict)
def list_stock(
    q: str = Query("", description="搜索试剂名称/品牌"),
    type: Optional[str] = Query(None),
    low_stock_only: bool = Query(False, description="只显示低于最低库存预警的"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    base = db.query(ReagentStock)
    if q.strip():
        kw = f"%{q.strip()}%"
        item_ids = [r[0] for r in db.query(ReagentItem.id).filter(
            or_(ReagentItem.name.like(kw), ReagentItem.brand.like(kw))
        ).all()]
        base = base.filter(ReagentStock.item_id.in_(item_ids))
    if type:
        item_ids2 = [r[0] for r in db.query(ReagentItem.id).filter(ReagentItem.type == type).all()]
        base = base.filter(ReagentStock.item_id.in_(item_ids2))
    if low_stock_only:
        # 仅显示低于最低库存预警的：join 试剂目录，按 min_stock>0 且 数量<min_stock 过滤
        base = base.join(ReagentItem, ReagentItem.id == ReagentStock.item_id).filter(
            ReagentItem.min_stock > 0,
            ReagentStock.quantity < ReagentItem.min_stock,
        )
    total = base.count()
    rows = base.order_by(ReagentStock.item_id).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": [ReagentStockRead.model_validate(r) for r in rows]}


# =============================================================================
# 3. 盘库
# =============================================================================

@router.get("/inventory-checks", response_model=dict)
def list_inventory_checks(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    total = db.query(InventoryCheck).count()
    rows = db.query(InventoryCheck).order_by(InventoryCheck.check_date.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": [InventoryCheckRead.model_validate(r) for r in rows]}


@router.get("/inventory-checks/{check_id}", response_model=InventoryCheckRead)
def get_inventory_check(check_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    check = db.query(InventoryCheck).get(check_id)
    if not check:
        raise HTTPException(404, "盘库未找到")
    return check


@router.post("/inventory-checks", response_model=InventoryCheckRead)
def create_inventory_check(
    data: InventoryCheckCreate, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    check = InventoryCheck(
        check_date=data.check_date,
        check_type=data.check_type,
        operator=user.full_name or user.username,
        remark=data.remark,
    )
    db.add(check)
    db.flush()  # 获取 id
    for it in data.items:
        item = InventoryCheckItem(
            check_id=check.id, item_id=it.item_id, batch_no=it.batch_no,
            expiry_date=it.expiry_date, recorded_quantity=it.recorded_quantity,
        )
        db.add(item)
        # 同步更新 reagent_stock
        stock = db.query(ReagentStock).filter(
            ReagentStock.item_id == it.item_id,
            ReagentStock.batch_no == it.batch_no,
        ).first()
        if stock:
            stock.quantity = it.recorded_quantity
            stock.last_updated = datetime.utcnow()
        else:
            db.add(ReagentStock(
                item_id=it.item_id, batch_no=it.batch_no,
                expiry_date=it.expiry_date, quantity=it.recorded_quantity,
            ))
    db.commit()
    db.refresh(check)
    return check


# =============================================================================
# 4. 订购
# =============================================================================

@router.get("/orders", response_model=dict)
def list_orders(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    total = db.query(ReagentOrder).count()
    rows = db.query(ReagentOrder).order_by(ReagentOrder.order_date.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": [ReagentOrderRead.model_validate(r) for r in rows]}


@router.get("/orders/{order_id}", response_model=ReagentOrderRead)
def get_order(order_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    order = db.query(ReagentOrder).get(order_id)
    if not order:
        raise HTTPException(404, "订购单未找到")
    return order


@router.post("/orders", response_model=ReagentOrderRead)
def create_order(
    data: ReagentOrderCreate, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    order = ReagentOrder(
        order_no=data.order_no, order_date=data.order_date,
        order_type=data.order_type, status="草稿",
        operator=user.full_name or user.username, remark=data.remark,
    )
    db.add(order)
    db.flush()
    for it in data.items:
        db.add(ReagentOrderItem(order_id=order.id, **it.model_dump()))
    db.commit()
    db.refresh(order)
    return order


@router.put("/orders/{order_id}", response_model=ReagentOrderRead)
def update_order(
    order_id: int, data: ReagentOrderCreate, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    order = db.query(ReagentOrder).get(order_id)
    if not order:
        raise HTTPException(404, "订购单未找到")
    order.order_date = data.order_date
    order.order_type = data.order_type
    order.status = data.status
    order.remark = data.remark
    order.operator = user.full_name or user.username
    # 先删旧细项再重建
    db.query(ReagentOrderItem).filter(ReagentOrderItem.order_id == order.id).delete()
    for it in data.items:
        db.add(ReagentOrderItem(order_id=order.id, **it.model_dump()))
    db.commit()
    db.refresh(order)
    return order


@router.delete("/orders/{order_id}")
def delete_order(
    order_id: int, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin")),
):
    order = db.query(ReagentOrder).get(order_id)
    if not order:
        raise HTTPException(404, "订购单未找到")
    db.delete(order)
    db.commit()
    return {"ok": True}


@router.get("/orders/{order_id}/export-form")
def export_order_form(
    order_id: int, db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """导出向设备科提交的订购表（含材料编码/名称/规格/品牌/数量/单位）。"""
    order = db.query(ReagentOrder).get(order_id)
    if not order:
        raise HTTPException(404, "订购单未找到")
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        raise HTTPException(400, "服务端未安装 openpyxl，无法导出 Excel")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "设备科订购表"
    headers = ["材料编码", "试剂名称", "规格", "品牌", "订购数量", "单位", "备注"]
    ws.append(headers)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = openpyxl.styles.PatternFill("solid", fgColor="1A365D")
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for it in order.items:
        item = db.query(ReagentItem).get(it.item_id)
        ws.append([
            item.material_code if item else "",
            item.name if item else f"(id={it.item_id})",
            item.spec if item else "",
            item.brand if item else "",
            it.ordered_quantity,
            item.unit if item else "",
            it.remark or "",
        ])
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = border
    widths = [16, 28, 22, 14, 12, 8, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"设备科订购表_{order.order_no}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# =============================================================================
# 5. 到货接收
# =============================================================================

@router.get("/receivings", response_model=dict)
def list_receivings(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    total = db.query(Receiving).count()
    rows = db.query(Receiving).order_by(Receiving.receipt_date.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": [ReceivingRead.model_validate(r) for r in rows]}


@router.get("/receivings/{receiving_id}", response_model=ReceivingRead)
def get_receiving(receiving_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    r = db.query(Receiving).get(receiving_id)
    if not r:
        raise HTTPException(404, "收货记录未找到")
    return r


@router.post("/receivings", response_model=ReceivingRead)
def create_receiving(
    data: ReceivingCreate, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    rec = Receiving(
        receipt_no=data.receipt_no, receipt_date=data.receipt_date,
        order_id=data.order_id, delivery_person=data.delivery_person,
        receiver=data.receiver or user.full_name or user.username, remark=data.remark,
    )
    db.add(rec)
    db.flush()
    for it in data.items:
        db.add(ReceivingItem(
            receiving_id=rec.id, item_id=it.item_id, batch_no=it.batch_no,
            expiry_date=it.expiry_date, quantity=it.quantity, remark=it.remark,
        ))
        # 库存增加
        stock = db.query(ReagentStock).filter(
            ReagentStock.item_id == it.item_id,
            ReagentStock.batch_no == it.batch_no,
        ).first()
        if stock:
            stock.quantity += it.quantity
            stock.last_updated = datetime.utcnow()
        else:
            db.add(ReagentStock(
                item_id=it.item_id, batch_no=it.batch_no,
                expiry_date=it.expiry_date, quantity=it.quantity,
            ))
    db.commit()
    db.refresh(rec)
    return rec


# =============================================================================
# 6. 月消耗
# =============================================================================

@router.get("/consumption", response_model=dict)
def list_consumption(
    year_month: Optional[str] = Query(None, description="YYYY-MM"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    base = db.query(ReagentConsumption)
    if year_month:
        base = base.filter(ReagentConsumption.year_month == year_month)
    total = base.count()
    rows = base.order_by(ReagentConsumption.year_month.desc(), ReagentConsumption.item_id).offset(
        (page - 1) * page_size
    ).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": [ReagentConsumptionRead.model_validate(r) for r in rows]}


@router.post("/consumption/_calculate")
def calculate_consumption(
    year_month: str = Query(..., description="YYYY-MM"),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    """盘库后触发：对所有 reagent_items 计算月消耗。

    公式：月消耗 = 期初库存(opening) + 本月入库(received) - 期末库存(closing)
      - 期末库存 = 当前实时库存合计（月末盘库已回写 ReagentStock）
      - 本月入库 = 当月到货接收合计（Receiving.receipt_date 在区间内）
      - 期初库存 = 上月消耗记录的期末库存
    """
    if not (len(year_month) == 7 and year_month[4] == "-"):
        raise HTTPException(400, "year_month 格式应为 YYYY-MM")
    try:
        y, m = int(year_month[:4]), int(year_month[5:7])
    except ValueError:
        raise HTTPException(400, "year_month 格式应为 YYYY-MM")
    month_start = date(y, m, 1)
    if m == 12:
        month_end = date(y + 1, 1, 1)
    else:
        month_end = date(y, m + 1, 1)

    items = db.query(ReagentItem).filter(ReagentItem.is_active == True).all()
    added, updated = 0, 0
    for it in items:
        # 期末库存：当前实时库存合计
        closing_qty = db.query(func.coalesce(func.sum(ReagentStock.quantity), 0)).filter(
            ReagentStock.item_id == it.id
        ).scalar() or 0
        # 本月入库：当月到货接收细项合计
        received_qty = db.query(func.coalesce(func.sum(ReceivingItem.quantity), 0)).join(
            Receiving, ReceivingItem.receiving_id == Receiving.id
        ).filter(
            ReceivingItem.item_id == it.id,
            Receiving.receipt_date >= month_start,
            Receiving.receipt_date < month_end,
        ).scalar() or 0
        # 期初库存：上月消耗记录的期末库存
        prev = db.query(ReagentConsumption).filter(
            ReagentConsumption.item_id == it.id,
        ).order_by(ReagentConsumption.year_month.desc()).first()
        opening = prev.closing_balance if prev else 0
        consumption = opening + received_qty - closing_qty
        if consumption < 0:
            consumption = 0

        existing = db.query(ReagentConsumption).filter(
            ReagentConsumption.item_id == it.id,
            ReagentConsumption.year_month == year_month,
        ).first()
        if existing:
            existing.opening_balance = opening
            existing.total_received = received_qty
            existing.closing_balance = closing_qty
            existing.consumption = consumption
            existing.calculated_at = datetime.utcnow()
            updated += 1
        elif opening or received_qty or closing_qty:
            # 仅当有数据时落库，避免大量空行
            db.add(ReagentConsumption(
                item_id=it.id, year_month=year_month,
                opening_balance=opening, total_received=received_qty,
                closing_balance=closing_qty, consumption=consumption,
            ))
            added += 1
    db.commit()
    return {"added": added, "updated": updated, "year_month": year_month}


# =============================================================================
# 7. Excel 导入试剂目录
# =============================================================================

@router.post("/items/_import-excel", response_model=ImportResult)
def import_reagent_from_excel(
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin")),
    file: UploadFile = File(...),
):
    """从 Excel 导入试剂目录（要求：第1行表头，需含 'name' 或 '试剂名称' 列）。"""
    result = ImportResult()
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(400, "服务端未安装 openpyxl，无法处理 Excel")
    wb = openpyxl.load_workbook(file.file)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Excel 为空")
    headers = [str(h or "").strip() for h in rows[0]]
    # 尝试推断列映射
    col_map = {}
    for col in ["name", "试剂名称", "材料名称", "名称", "名称(试剂名称)", "试剂名"]:
        if col in headers:
            col_map["name"] = headers.index(col)
            break
    if "name" not in col_map:
        raise HTTPException(400, "找不到 'name/试剂名称' 列")
    for col in ["type", "类型", "试剂类型"]:
        if col in headers:
            col_map["type"] = headers.index(col)
            break
    for col in ["category", "类别", "专业", "专业组"]:
        if col in headers:
            col_map["category"] = headers.index(col)
            break
    for col in ["library", "责任库", "库"]:
        if col in headers:
            col_map["library"] = headers.index(col)
            break
    for col in ["brand", "品牌", "生产厂家", "厂家"]:
        if col in headers:
            col_map["brand"] = headers.index(col)
            break
    for col in ["spec", "规格", "规格型号", "包装规格"]:
        if col in headers:
            col_map["spec"] = headers.index(col)
            break
    for col in ["material_code", "材料编码", "编码", "订购编码"]:
        if col in headers:
            col_map["material_code"] = headers.index(col)
            break
    for col in ["unit", "单位"]:
        if col in headers:
            col_map["unit"] = headers.index(col)
            break
    for col in ["manufacturer", "生产厂家2", "制造商"]:
        if col in headers:
            col_map["manufacturer"] = headers.index(col)
            break
    for col in ["supplier", "供应商"]:
        if col in headers:
            col_map["supplier"] = headers.index(col)
            break
    for col in ["unit_price", "单价", "价格", "参考单价"]:
        if col in headers:
            col_map["unit_price"] = headers.index(col)
            break
    for col in ["remark", "备注", "说明"]:
        if col in headers:
            col_map["remark"] = headers.index(col)
            break

    imported = 0
    skipped = 0
    for i, row in enumerate(rows[1:], 2):
        try:
            name = str(row[col_map["name"]] or "").strip()
            if not name:
                skipped += 1
                continue
            # 主去重键：材料编码优先，回退名称
            code = str(row[col_map["material_code"]]).strip() if "material_code" in col_map else ""
            existing = None
            if code:
                existing = db.query(ReagentItem).filter(ReagentItem.material_code == code).first()
            if existing is None:
                existing = db.query(ReagentItem).filter(ReagentItem.name == name).first()
            if existing:
                # 已存在：用本次导入信息补全/覆盖目录字段
                for key in ("material_code", "spec", "brand", "unit", "category", "library",
                            "type", "unit_price", "manufacturer", "supplier", "remark"):
                    if key not in col_map:
                        continue
                    v = row[col_map[key]]
                    if v is None or str(v).strip() == "":
                        continue
                    if key == "unit_price":
                        try:
                            setattr(existing, key, Decimal(str(v)))
                        except Exception:
                            pass
                    elif key == "category":
                        val = str(v).strip()
                        setattr(existing, key, val)
                        if "library" not in col_map:
                            setattr(existing, "library", derive_library(val))
                    elif key == "library":
                        setattr(existing, key, str(v).strip())
                    elif key == "type":
                        setattr(existing, key, str(v).strip())
                    else:
                        setattr(existing, key, str(v).strip())
                skipped += 1
                continue
            item = ReagentItem(name=name)
            if "type" in col_map:
                item.type = str(row[col_map["type"]] or "试剂").strip()
            else:
                item.type = detect_reagent_type(name)  # 按名称自动识别
            if "category" in col_map:
                item.category = str(row[col_map["category"]] or "").strip()
            if "library" in col_map:
                item.library = str(row[col_map["library"]] or "").strip()
            elif item.category:
                item.library = derive_library(item.category)  # 由专业组推导责任库
            if "brand" in col_map:
                item.brand = str(row[col_map["brand"]] or "").strip()
            if "spec" in col_map:
                item.spec = str(row[col_map["spec"]] or "").strip()
            if "material_code" in col_map:
                item.material_code = code
            if "unit" in col_map:
                item.unit = str(row[col_map["unit"]] or "").strip()
            if "manufacturer" in col_map:
                item.manufacturer = str(row[col_map["manufacturer"]] or "").strip()
            if "supplier" in col_map:
                item.supplier = str(row[col_map["supplier"]] or "").strip()
            if "remark" in col_map:
                item.remark = str(row[col_map["remark"]] or "").strip()
            if "unit_price" in col_map:
                v = row[col_map["unit_price"]]
                if v is not None and str(v).strip() != "":
                    try:
                        item.unit_price = Decimal(str(v))
                    except Exception:
                        pass
            db.add(item)
            imported += 1
        except Exception as e:
            result.errors.append(f"第{i}行: {e}")
            skipped += 1
    db.commit()
    result.total = len(rows) - 1
    result.imported = imported
    result.skipped = skipped
    return result


# =============================================================================
# 8. 项目 ↔ 试剂 / 仪器 ↔ 耗材 关联管理
# =============================================================================

import re as _re


def _norm_match(s: str) -> str:
    """归一化用于模糊匹配：仅去空白/标点与 R1/R2 标记，保留字母数字与中文
    （刻意保留型号代码 DxI800 / ACL / cobas 等，以便耗材能按仪器型号匹配）。"""
    s = str(s or "")
    s = _re.sub(r"[Rr][1-9]\b", "", s)                # 去 R1/R2 标记（避免干扰）
    s = _re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]", "", s)  # 仅留字母数字与中文（保留型号代码）
    return s.strip().lower()


@router.post("/associations/_auto-match", response_model=dict)
def auto_match_associations(
    reset: bool = Query(False, description="true 时先清空已有自动匹配记录再重新生成"),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    """根据名称模糊匹配，自动生成 项目↔试剂 与 耗材↔仪器 关联。

    - 试剂/校准品/质控品：名称 vs 项目 name/aliases 模糊匹配，role 取试剂类型(试剂/校准品/质控品)。
    - 耗材：名称 vs 仪器 name/model 关键词匹配，role=耗材。
    返回新增条数与无法自动匹配（需人工审核）的清单。
    """
    if reset:
        db.query(TestItemReagent).delete()
        db.query(InstrumentReagent).delete()

    tests = db.query(TestItem).all()
    test_norms = [
        (t.id, _norm_match(t.name),
         [_norm_match(a) for a in str(t.aliases or "").split(",") if a.strip()])
        for t in tests
    ]
    reagents = db.query(ReagentItem).all()

    tir_added, tir_unmatched = 0, []
    for r in reagents:
        if r.type == "耗材":
            continue
        rn = _norm_match(r.name)
        best, best_score = None, 0
        for tid, tn, tals in test_norms:
            for cand in ([tn] + tals):
                if not cand:
                    continue
                score = len(cand) if cand in rn else (len(rn) if (rn and rn in cand) else 0)
                if score > best_score:
                    best_score, best = score, tid
        if best and best_score >= 2:
            if not db.query(TestItemReagent).filter_by(test_item_id=best, reagent_item_id=r.id).first():
                db.add(TestItemReagent(
                    test_item_id=best, reagent_item_id=r.id,
                    role=r.type if r.type in ("校准品", "质控品") else "试剂",
                    auto_matched=True,
                ))
                tir_added += 1
        else:
            tir_unmatched.append({"reagent_id": r.id, "name": r.name, "type": r.type})

    insts = db.query(Instrument).all()
    inst_norms = [
        (ins.id, _norm_match(ins.name) + _norm_match(getattr(ins, "model", None) or ""))
        for ins in insts
    ]
    ir_added, ir_unmatched = 0, []
    for r in reagents:
        if r.type != "耗材":
            continue
        rn = _norm_match(r.name)
        best, best_score = None, 0
        for iid, inorm in inst_norms:
            if not inorm:
                continue
            score = len(inorm) if inorm in rn else (len(rn) if (rn and rn in inorm) else 0)
            if score > best_score:
                best_score, best = score, iid
        if best and best_score >= 3:
            if not db.query(InstrumentReagent).filter_by(instrument_id=best, reagent_item_id=r.id).first():
                db.add(InstrumentReagent(instrument_id=best, reagent_item_id=r.id, role="耗材", auto_matched=True))
                ir_added += 1
        else:
            ir_unmatched.append({"reagent_id": r.id, "name": r.name})

    db.commit()
    return {
        "test_item_reagents_added": tir_added,
        "instrument_reagents_added": ir_added,
        "test_item_unmatched": tir_unmatched,
        "instrument_unmatched": ir_unmatched,
    }


def _enrich_test_item_reagents(rows):
    out = []
    for r in rows:
        ti = r.test_item
        ri = r.reagent_item
        out.append({
            "id": r.id, "test_item_id": r.test_item_id, "reagent_item_id": r.reagent_item_id,
            "role": r.role, "auto_matched": r.auto_matched, "remark": r.remark,
            "test_item_name": ti.name if ti else "",
            "reagent_name": ri.name if ri else "", "reagent_type": ri.type if ri else "",
            "reagent_library": ri.library if ri else "",
        })
    return out


def _enrich_instrument_reagents(rows):
    out = []
    for r in rows:
        ins = r.instrument
        ri = r.reagent_item
        out.append({
            "id": r.id, "instrument_id": r.instrument_id, "reagent_item_id": r.reagent_item_id,
            "role": r.role, "auto_matched": r.auto_matched, "remark": r.remark,
            "instrument_name": ins.name if ins else "",
            "reagent_name": ri.name if ri else "", "reagent_type": ri.type if ri else "",
            "reagent_library": ri.library if ri else "",
        })
    return out


@router.get("/associations/test-items", response_model=dict)
def list_test_item_reagents(
    test_item_id: Optional[int] = Query(None),
    reagent_id: Optional[int] = Query(None),
    auto_only: Optional[bool] = Query(None),
    q: str = Query("", description="搜索项目名/试剂名"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    base = db.query(TestItemReagent)
    if test_item_id:
        base = base.filter(TestItemReagent.test_item_id == test_item_id)
    if reagent_id:
        base = base.filter(TestItemReagent.reagent_item_id == reagent_id)
    if auto_only is not None:
        base = base.filter(TestItemReagent.auto_matched == auto_only)
    if q.strip():
        kw = f"%{q.strip()}%"
        base = base.join(ReagentItem, ReagentItem.id == TestItemReagent.reagent_item_id).filter(
            or_(TestItemReagent.test_item_id.in_(
                [t.id for t in db.query(TestItem).filter(TestItem.name.like(kw)).all()]),
                ReagentItem.name.like(kw))
        )
    total = base.count()
    rows = base.order_by(TestItemReagent.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": _enrich_test_item_reagents(rows)}


@router.post("/associations/test-items", response_model=TestItemReagentRead)
def create_test_item_reagent(
    data: TestItemReagentCreate, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    if db.query(TestItemReagent).filter_by(test_item_id=data.test_item_id, reagent_item_id=data.reagent_item_id).first():
        raise HTTPException(409, "该关联已存在")
    obj = TestItemReagent(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/associations/test-items/{rel_id}")
def delete_test_item_reagent(
    rel_id: int, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    obj = db.query(TestItemReagent).get(rel_id)
    if not obj:
        raise HTTPException(404, "关联未找到")
    db.delete(obj)
    db.commit()
    return {"ok": True}


@router.get("/associations/instruments", response_model=dict)
def list_instrument_reagents(
    instrument_id: Optional[int] = Query(None),
    reagent_id: Optional[int] = Query(None),
    auto_only: Optional[bool] = Query(None),
    q: str = Query("", description="搜索仪器名/试剂名"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    base = db.query(InstrumentReagent)
    if instrument_id:
        base = base.filter(InstrumentReagent.instrument_id == instrument_id)
    if reagent_id:
        base = base.filter(InstrumentReagent.reagent_item_id == reagent_id)
    if auto_only is not None:
        base = base.filter(InstrumentReagent.auto_matched == auto_only)
    if q.strip():
        kw = f"%{q.strip()}%"
        base = base.join(ReagentItem, ReagentItem.id == InstrumentReagent.reagent_item_id).filter(
            or_(InstrumentReagent.instrument_id.in_(
                [i.id for i in db.query(Instrument).filter(Instrument.name.like(kw)).all()]),
                ReagentItem.name.like(kw))
        )
    total = base.count()
    rows = base.order_by(InstrumentReagent.id).offset((page - 1) * page_size).limit(page_size).all()
    return {"total": total, "page": page, "page_size": page_size,
            "items": _enrich_instrument_reagents(rows)}


@router.post("/associations/instruments", response_model=InstrumentReagentRead)
def create_instrument_reagent(
    data: InstrumentReagentCreate, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    if db.query(InstrumentReagent).filter_by(instrument_id=data.instrument_id, reagent_item_id=data.reagent_item_id).first():
        raise HTTPException(409, "该关联已存在")
    obj = InstrumentReagent(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/associations/instruments/{rel_id}")
def delete_instrument_reagent(
    rel_id: int, db: Session = Depends(get_db),
    user: User = Depends(require_roles("admin", "lab_technician")),
):
    obj = db.query(InstrumentReagent).get(rel_id)
    if not obj:
        raise HTTPException(404, "关联未找到")
    db.delete(obj)
    db.commit()
    return {"ok": True}
