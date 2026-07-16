"""室间质评（EQA）接口：年度计划 CRUD、检测提醒、半年/年度总结统计与文字、质评报告导入。"""
import os
import re
import json
import shutil
import fitz  # PyMuPDF，离线抽取 PDF 文本
from datetime import datetime
from io import BytesIO
from pathlib import Path

from fastapi import Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from ...core.crud_base import make_router
from ...core.config import DATA_DIR
from ...core.database import get_db
from ...core.security import get_current_user
from ...models.eqa import EqaPlan, EqaSummary
from ...models.user import User
from ...models.test_item import TestItem
from ...schemas import (
    EqaPlanCreate,
    EqaPlanRead,
    EqaPlanUpdate,
    EqaSummaryRead,
    EqaSummaryUpdate,
)
from ...services.notification_service import compute_eqa_alerts, refresh_eqa_notifications


def _after_eqa_write(db, action, obj):
    """CRUD 后刷新室间质评回报提醒（忽略 action/obj 参数）。"""
    refresh_eqa_notifications(db)


router = make_router(
    EqaPlan,
    EqaPlanRead,
    EqaPlanCreate,
    EqaPlanUpdate,
    search_fields=["org", "program", "item", "round_no"],
    filter_fields=["year", "org", "program", "group", "returned", "qualified"],
    prefix="/eqa-plans",
    after_write=_after_eqa_write,
    write_roles=("admin", "qc_manager"),
)


# ---------------------------------------------------------------------------
# 检测提醒：基于未上报计划的上报截止日期（供前端"检测提醒"卡片与首页提醒中心）
# ---------------------------------------------------------------------------
@router.get("/alerts")
def eqa_alerts(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """返回所有未回报且即将到期（或已逾期）的质评计划提醒列表。"""
    return compute_eqa_alerts(db)


# ---------------------------------------------------------------------------
# 半年/年度总结统计
# ---------------------------------------------------------------------------
_H1 = ["01", "02", "03", "04", "05", "06"]
_H2 = ["07", "08", "09", "10", "11", "12"]


@router.get("/summary")
def eqa_summary(
    year: int = 0,
    half: int = 0,  # 0=全年, 1=上半年(1-6月), 2=下半年(7-12月)
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """某年（及指定半年）的室间质评统计：计划数、已回报、合格率等 + 每轮明细。"""
    q = db.query(EqaPlan).filter(EqaPlan.year == year)
    if half == 1:
        q = q.filter(func.substr(EqaPlan.due_date, 6, 2).in_(_H1))
    elif half == 2:
        q = q.filter(func.substr(EqaPlan.due_date, 6, 2).in_(_H2))
    plans = q.order_by(EqaPlan.due_date, EqaPlan.org, EqaPlan.program).all()

    total = len(plans)
    returned = sum(1 for p in plans if p.returned)
    scored = [p for p in plans if (p.result or p.score or p.qualified)]
    qualified_count = sum(1 for p in scored if p.qualified)
    unqualified = len(scored) - qualified_count

    return {
        "year": year,
        "half": half,
        "total": total,
        "returned": returned,
        "not_returned": total - returned,
        "return_rate": round(returned / total * 100, 1) if total else 0.0,
        "scored": len(scored),
        "qualified": qualified_count,
        "unqualified": unqualified,
        "qualify_rate": round(qualified_count / len(scored) * 100, 1) if scored else 0.0,
        "plans": [EqaPlanRead.model_validate(p).model_dump() for p in plans],
    }


# ---------------------------------------------------------------------------
# 半年/年度总结：按专业分类（生化+凝血 / 免疫）的细项合格率统计
# ---------------------------------------------------------------------------
_NE_RE = re.compile(r"^(成绩不适用|不适用|不评价)")


def _category_of(group: str) -> str:
    g = (group or "").strip()
    if g in ("生化", "凝血"):
        return "生化+凝血"
    if g == "免疫":
        return "免疫"
    return "其他"


def _is_not_evaluated(result: str) -> bool:
    return bool(_NE_RE.match((result or "").strip()))


# 总结分类：拆成两份独立报告（两人分别负责），「其他」不进任何报告
_SUMMARY_CATEGORIES = ["生化+凝血", "免疫"]
_CATEGORY_SLUG = {"生化+凝血": "BC", "免疫": "IM"}

# 质评部门（org）显示名与文件名 slug 映射
_ORG_DISPLAY = {"卫健委": "国家卫健委临检中心", "北京市": "北京市临检中心"}
_ORG_SLUG = {"卫健委": "NCCL", "北京市": "BJ"}


def _normalize_category(category: str | None) -> str:
    """将传入分类归一化为受支持的两类之一；缺省/非法 → 生化+凝血。"""
    c = (category or "").strip()
    return c if c in _SUMMARY_CATEGORIES else "生化+凝血"


def _normalize_department(department: str | None) -> str:
    """将传入质评部门归一化为 org 值（卫健委/北京市）；接受显示名或 org。"""
    d = (department or "").strip()
    if d in _ORG_DISPLAY:
        return d
    for k, v in _ORG_DISPLAY.items():
        if d == v:
            return k
    return "卫健委"


def _compute_summary_by_category(year: int, half: int, db: Session, category: str | None = None, department: str | None = None) -> dict:
    """按（专业组 × 质评部门）统计细项合格率；细项按（标本 + 分析物）去重（多项目组同名算 1 项）。

    category：生化+凝血 / 免疫（指定时只统计该组）；department：卫健委 / 北京市
    （指定时只统计该质评部门）。「其他」类别与未指定的部门之外的 org 不纳入。
    标本来自项目组（尿液→尿液、脑脊液→脑脊液、血气→血气、凝血类→血浆，其余→血清）；
    token 自带（尿液）/（脑脊液）等后缀时优先采用。去重键 = 标本 + 分析物，使血清/尿液/
    脑脊液/血气同名项分别计数（不再错误合并）。「正确度验证」类项目组不计入常规质评细项。
    每条唯一细项继承其评价状态：全部评价为合格→合格；任一为不合格→不合格；
    全部为不评价→不评价；无任何评价记录（未回报）→仅计入细项总数，不计入评价分母。
    """
    cat_target = _normalize_category(category) if category else None
    dept_target = _normalize_department(department) if department else None

    q = db.query(EqaPlan).filter(EqaPlan.year == year)
    if half == 1:
        q = q.filter(func.substr(EqaPlan.due_date, 6, 2).in_(_H1))
    elif half == 2:
        q = q.filter(func.substr(EqaPlan.due_date, 6, 2).in_(_H2))
    if dept_target:
        q = q.filter(EqaPlan.org == dept_target)
    plans = q.order_by(EqaPlan.due_date, EqaPlan.org, EqaPlan.program).all()
    # 只保留纳入报告的分类；指定 category 时进一步收窄
    plans = [p for p in plans if _category_of(p.group) in _SUMMARY_CATEGORIES
             and (cat_target is None or _category_of(p.group) == cat_target)]

    # 按分类收集：项目组（去重集合）+ 细项目（按名称去重，记录评价状态）
    cats: dict = {}

    def bucket(name):
        if name not in cats:
            cats[name] = {
                "category": name,
                "programs": set(),
                "items": {},  # item_name -> {"status": [...], "programs": set(), "plan": EqaPlan}
            }
        return cats[name]

    for p in plans:
        # 正确度验证类项目不计为常规质评细项（仅方法学正确度考核，不参与合格率统计）
        if "正确度" in (p.program or ""):
            continue
        b = bucket(_category_of(p.group))
        b["programs"].add(p.program)
        spec_plan = _specimen_of_program(p.program)
        subs = _split_items(p.item)
        if not subs:
            subs = ["（未填细项）"]
        if _is_not_evaluated(p.result):
            st = "NE"
        elif p.qualified:
            st = "Q"
        elif p.result or p.score or p.qualified:
            st = "U"
        else:
            st = None  # 未回报
        for s in subs:
            spec, an = _split_specimen(s, spec_plan)
            key = f"{spec}::{an}"
            rec = b["items"].setdefault(key, {"status": [], "programs": set(),
                                              "plan": p, "label": an, "specimen": spec})
            rec["programs"].add(p.program)
            if st is not None:
                rec["status"].append(st)

    order = [cat_target] if cat_target else _SUMMARY_CATEGORIES
    rows = []
    for name in order:
        if name not in cats:
            continue
        b = cats[name]
        programs = len(b["programs"])
        items_total = len(b["items"])
        qualified = unqualified = not_evaluated = items_evaluated = 0
        unqualified_list = []
        not_evaluated_list = []
        for item_name, rec in b["items"].items():
            statuses = rec["status"]
            if not statuses:
                continue  # 未回报：仅计入细项总数
            items_evaluated += 1
            example = rec["plan"]
            an = rec.get("label", item_name)
            spec = rec.get("specimen", "血清")
            display = an if spec == "血清" else f"{an}（{spec}）"
            if all(x == "NE" for x in statuses):
                not_evaluated += 1
                not_evaluated_list.append({
                    "program": example.program, "item": display, "group": example.group,
                    "result": example.result, "score": example.score,
                })
            elif "U" in statuses:
                unqualified += 1
                unqualified_list.append({
                    "program": example.program, "item": display, "group": example.group,
                    "result": example.result, "score": example.score,
                })
            else:
                qualified += 1
        denom = qualified + unqualified
        rate = round(qualified / denom * 100, 1) if denom else None
        rows.append({
            "category": name,
            "programs": programs,
            "items_total": items_total,
            "items_evaluated": items_evaluated,
            "qualified": qualified,
            "unqualified": unqualified,
            "not_evaluated": not_evaluated,
            "qualify_rate": rate,
            "unqualified_list": unqualified_list,
            "not_evaluated_list": not_evaluated_list,
        })

    tot = {"programs": 0, "items_total": 0, "items_evaluated": 0,
           "qualified": 0, "unqualified": 0, "not_evaluated": 0}
    if len(rows) == 1:
        # 单分类（UI 常规路径）：合计即该分类本身
        r = rows[0]
        tot = {
            "programs": r["programs"], "items_total": r["items_total"],
            "items_evaluated": r["items_evaluated"], "qualified": r["qualified"],
            "unqualified": r["unqualified"], "not_evaluated": r["not_evaluated"],
            "qualify_rate": r["qualify_rate"],
        }
    else:
        prog_union = set()
        for b in cats.values():
            prog_union |= b["programs"]
            tot["items_total"] += len(b["items"])
        # 跨分类去重评价状态（生化+凝血与免疫同名细项极罕见，仍做一致处理）
        union_items = {}
        for b in cats.values():
            for item_name, rec in b["items"].items():
                u = union_items.setdefault(item_name, {"status": []})
                u["status"].extend(rec["status"])
        for item_name, rec in union_items.items():
            statuses = rec["status"]
            if not statuses:
                continue
            items_evaluated += 1
            if all(x == "NE" for x in statuses):
                not_evaluated += 1
            elif "U" in statuses:
                unqualified += 1
            else:
                qualified += 1
        tot["programs"] = len(prog_union)
        tot["items_evaluated"] = items_evaluated
        tot["qualified"] = qualified
        tot["unqualified"] = unqualified
        tot["not_evaluated"] = not_evaluated
        tdenom = qualified + unqualified
        tot["qualify_rate"] = round(qualified / tdenom * 100, 1) if tdenom else None

    return {
        "year": year, "half": half,
        "department": dept_target, "category": cat_target,
        "categories": rows, "total": tot, "plan_count": len(plans),
    }


@router.get("/summary-by-category")
def eqa_summary_by_category(
    year: int = 0,
    half: int = 0,
    category: str = "",
    department: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """按（专业组 × 质评部门）的细项合格率统计（含不合格/不评价明细，细项已去重）。

    传 category（生化+凝血 / 免疫）+ department（卫健委 / 北京市）时只统计该组合，
    用于两人分别、按部门出独立报告。
    """
    return _compute_summary_by_category(year, half, db, category or None, department or None)


# 总结 Word 存档目录
EQA_SUMMARY_DIR = DATA_DIR / "eqa_summaries"
os.makedirs(EQA_SUMMARY_DIR, exist_ok=True)


class EqaSummaryGenerate(BaseModel):
    year: int
    half: int = 0
    department: str = "卫健委"
    category: str = "生化+凝血"
    summary_text: str = ""


@router.post("/summary-generate")
def generate_eqa_summary(
    body: EqaSummaryGenerate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """保存某（部门×专业组）总结文字并生成 BG-SM-CZ-020 格式 Word，自动存档；返回下载地址与统计。"""
    dept = _normalize_department(body.department)
    cat = _normalize_category(body.category)
    stats = _compute_summary_by_category(body.year, body.half, db, cat, dept)
    slug = f"{_ORG_SLUG.get(dept, 'ORG')}_{_CATEGORY_SLUG.get(cat, 'BC')}"
    fname = f"{body.year}_H{body.half}_{slug}.docx"
    out_path = EQA_SUMMARY_DIR / fname
    from ...services.eqa_summary_doc import build_eqa_summary_doc
    build_eqa_summary_doc(body.year, body.half, stats, body.summary_text or "", str(out_path),
                          category=cat, department=dept)

    rep = db.query(EqaSummary).filter_by(year=body.year, half=body.half, department=dept, category=cat).first()
    if not rep:
        rep = EqaSummary()
        db.add(rep)
    rep.year = body.year
    rep.half = body.half
    rep.department = dept
    rep.category = cat
    rep.summary_text = body.summary_text or ""
    rep.docx_path = fname
    rep.generated_at = datetime.utcnow()
    db.commit()
    db.refresh(rep)
    from urllib.parse import quote
    return {
        "id": rep.id,
        "year": rep.year,
        "half": rep.half,
        "department": rep.department,
        "category": rep.category,
        "docx_path": rep.docx_path,
        "docx_url": f"/api/v1/eqa-plans/summary-docx?year={rep.year}&half={rep.half}"
                    f"&department={quote(rep.department)}&category={quote(rep.category)}",
        "generated_at": rep.generated_at,
        "stats": stats,
    }


@router.get("/summary-docx")
def download_eqa_summary_doc(
    year: int = 0,
    half: int = 0,
    department: str = "卫健委",
    category: str = "生化+凝血",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """下载已生成的某（部门×专业组）总结 Word（自动存档文件）。"""
    dept = _normalize_department(department)
    cat = _normalize_category(category)
    rep = db.query(EqaSummary).filter_by(year=year, half=half, department=dept, category=cat).first()
    if not rep or not rep.docx_path:
        raise HTTPException(status_code=404, detail="尚未生成总结报告，请先保存并生成")
    path = EQA_SUMMARY_DIR / rep.docx_path
    if not path.exists():
        raise HTTPException(status_code=404, detail="报告文件缺失")
    return FileResponse(
        str(path),
        filename=rep.docx_path,
        headers={"Cache-Control": "no-store"},
    )


# ---------------------------------------------------------------------------
# 半年总结文字（可编辑）
# ---------------------------------------------------------------------------
@router.get("/summary-text", response_model=EqaSummaryRead)
def get_eqa_summary_text(
    year: int = 0,
    half: int = 1,
    department: str = "卫健委",
    category: str = "生化+凝血",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    dept = _normalize_department(department)
    cat = _normalize_category(category)
    rep = db.query(EqaSummary).filter_by(year=year, half=half, department=dept, category=cat).first()
    if not rep:
        rep = EqaSummary(year=year, half=half, department=dept, category=cat)
        db.add(rep)
        db.commit()
        db.refresh(rep)
    return rep


@router.put("/summary-text", response_model=EqaSummaryRead)
def upsert_eqa_summary_text(
    body: EqaSummaryUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    dept = _normalize_department(body.department)
    cat = _normalize_category(body.category)
    rep = db.query(EqaSummary).filter_by(year=body.year, half=body.half, department=dept, category=cat).first()
    if not rep:
        rep = EqaSummary()
        db.add(rep)
    rep.year = body.year
    rep.half = body.half
    rep.department = dept
    rep.category = cat
    rep.summary_text = body.summary_text or ""
    db.commit()
    db.refresh(rep)
    return rep


# ---------------------------------------------------------------------------
# 质评报告导入 / 下载 / 删除（卫健委、北京市 PDF 质评报告留档「口子」）
# ---------------------------------------------------------------------------
# 报告存放目录统一走 DATA_DIR（线上由 ENV DATA_DIR=/app/data 指向 CFS 持久卷；
# 早期曾硬编码 parents[3]/data/eqa_reports，在容器里虽恰好是 /app/data/eqa_reports，
# 但本地却落在 backend/data/eqa_reports，与 uploads/documents 等模块不一致。
# 统一用 DATA_DIR 后，本地(data/eqa_reports)与线上(/app/data/eqa_reports)路径一致且均在持久卷上。）
EQA_REPORT_DIR = DATA_DIR / "eqa_reports"
os.makedirs(EQA_REPORT_DIR, exist_ok=True)


# 非项目名行：报告标题/页眉页脚/表头/单位等，用于定位「项目名」时跳过
_NONAME = re.compile(
    r"室间质量评价|实验室编码|实验室名称|测定日期|统计日期|报表打印日期|打印日期|"
    r"统计结果|国家卫生|卫生健康委|第\d+次|页|"
    r"你室结果|所有实验室稳健均值|允许范围|下限靶值上限评价结果|样本编号|偏倚|"
    r"所属组|方法|仪器|试剂|本组实验室数|校准物|建议|项目：|"
    r"成绩汇总|成绩解释|成绩说明|科室|"
    r"成功|当前解释|及格数|总数|得分%|解释|评价结果|成绩|汇总"
)


def _is_name_line(l: str) -> bool:
    """判断一行是否可能是 analyte 项目名（而非表头/标题/单位/数字/日期）。"""
    l = l.strip()
    if not l:
        return False
    if "%" in l:
        return False
    if re.fullmatch(r"[\d.]+", l):
        return False
    if re.fullmatch(r"[\d\s\-]+", l):  # 纯数字/空格/短横（如页码 "4 - 0"）
        return False
    if re.fullmatch(r"[A-Za-z/μ·]+", l):  # 单位（mmol/L, nmol/L, g/L…）
        return False
    if re.search(r"^\d{4}[-/年]", l):  # 日期
        return False
    if _NONAME.search(l):
        return False
    return True


def _extract_nhc_blocks(lines):
    """卫健委报告：按「成绩」锚点切块，返回 [(项目名, 成绩str|None, is_na)]。

    卫健委每项结构：项目名 + 表头行 + 若干样本行(末列 通过/不适用/不评价) + 末行「成绩NN％/成绩不适用」。
    项目名 = 块内第一条「非表头/非标题/非单位/非数字」的行（即 analyte 名本身），
    能正确跳过「允许范围」等表格列头，不受 item 别名录入差异影响。
    """
    anchors = []  # (line_index, score_or_None, is_na)
    for i, l in enumerate(lines):
        # 排除「成绩汇总/成绩解释」等标题行，仅匹配真正的成绩锚点
        if l.startswith("成绩") and not re.search(r"汇总|解释|说明|备注|标题", l):
            is_na = bool(re.search(r"不适用|不评价", l))
            m = re.search(r"(\d{1,3})\s*[％%]", l)
            if not m and i + 1 < len(lines):
                m = re.search(r"(\d{1,3})\s*[％%]", lines[i + 1])
            score = m.group(1) if m else None
            anchors.append((i, score, is_na))
    if not anchors:
        return []
    blocks = []
    for k, (ai, score, is_na) in enumerate(anchors):
        start = anchors[k - 1][0] + 1 if k > 0 else 0
        # 项目名 = 块内自 start 起第一条「像项目名」的行
        name = None
        for j in range(start, ai):
            if _is_name_line(lines[j]):
                name = lines[j].strip()
                break
        blocks.append((name, score, is_na))
    return blocks


def _extract_bj_lines(lines):
    """北京市报告：按「满意/不满意」锚点切块，返回 [(项目名, satisfied, score_pct|None)]。

    北京每项结构：项目名 + 及格数/总数/得分%(三数字行) + 「满意/不满意」。
    项目名 = 块内第一条「像项目名」的行；得分% = 满意前最后一个纯数字行。
    """
    anchors = []  # (line_index, is_unsat)
    for i, l in enumerate(lines):
        if l == "满意" or l == "不满意" or re.match(r"^(满意|不满意)$", l):
            anchors.append((i, l == "不满意"))
    if not anchors:
        return []
    blocks = []
    for k, (ai, unsat) in enumerate(anchors):
        start = anchors[k - 1][0] + 1 if k > 0 else 0
        name = None
        for j in range(start, ai):
            if _is_name_line(lines[j]):
                name = lines[j].strip()
                break
        # 得分% = 满意/不满意 前最后一个纯数字行
        score_pct = None
        for j in range(ai - 1, start - 1, -1):
            bl = lines[j].strip()
            if re.fullmatch(r"\d+", bl):
                score_pct = bl
                break
        blocks.append((name, not unsat, score_pct))
    return blocks


def parse_eqa_report(pdf_path: str, item_text: str = None) -> dict:
    """解析质评报告 PDF，抽取成绩/合格状态/得分，并「点名」列出不适用或非满分的子项。

    从报告正文权威解析（不依赖 item 别名的录入差异），覆盖两类报告：
      - 卫健委：逐项分块「项目名 + 通过×N + 成绩NN％」；不适用子项为「项目名 + 不适用×N + 成绩不适用」
      - 北京市：逐项目「项目名 + 及格数/总数/得分% + 满意/不满意」
    返回 {result, qualified(bool|None), score(str), confidence, evidence}。
    result 会显式点名异常子项，例如：
      「成绩100%（脂蛋白(a)不适用）」「成绩未达标（总钙92分）」「不合格（钾、钠不满意）」
    若整体合格但仅部分子项不适用，仍标合格并备注不适用子项，避免误判为不合格。
    """
    try:
        doc = fitz.open(pdf_path)
        txt = "\n".join(pg.get_text() for pg in doc)
    except Exception:
        return {}
    if not txt.strip():
        return {}
    lines = [l.strip() for l in txt.splitlines()]
    evidence = re.sub(r"\s+", " ", txt[:80])

    # 卫健委：逐项切块
    nhc = [b for b in _extract_nhc_blocks(lines) if b[0]]
    # 北京市：逐项目
    bj = [b for b in _extract_bj_lines(lines) if b[0]]

    if nhc:
        na_items = [n for n, s, na in nhc if na]
        non100 = [(n, s) for n, s, na in nhc if s is not None and s != "100"]
        if na_items and not non100 and len(na_items) < len(nhc):
            # 整体合格，仅部分子项不适用 -> 备注点名，不误判不合格
            note = "（" + "、".join(f"{n}不适用" for n in na_items) + "）"
            return {"result": "成绩100%" + note, "qualified": True, "score": "100",
                    "confidence": "high", "evidence": evidence}
        if non100:
            # 有子项成绩非100 -> 该子项未达标，整体不合格并点名
            notes = [f"{n}不适用" for n in na_items] + [f"{n}({s}分)" for n, s in non100]
            note = "（" + "、".join(notes) + "）"
            return {"result": "成绩未达标" + note, "qualified": False, "score": non100[0][1],
                    "confidence": "high", "evidence": evidence}
        if na_items and len(na_items) == len(nhc):
            return {"result": "成绩不适用(不予评价)", "qualified": None, "score": "",
                    "confidence": "high", "evidence": evidence}
        return {"result": "成绩100%", "qualified": True, "score": "100",
                "confidence": "high", "evidence": evidence}

    if bj:
        unsat = [n for n, sat, _ in bj if sat is False]
        non100 = [(n, s) for n, sat, s in bj if s is not None and s != "100"]
        if unsat and not non100:
            note = "（" + "、".join(f"{n}不满意" for n in unsat) + "）"
            return {"result": "不合格" + note, "qualified": False, "score": "",
                    "confidence": "high", "evidence": evidence}
        if non100 and not unsat:
            # 整体满意(合格)但存在得分%<100 的子项 -> 标合格并点名非满分项目
            note = "（" + "、".join(f"{n}({s}分)" for n, s in non100) + "）"
            return {"result": "合格" + note, "qualified": True, "score": "",
                    "confidence": "high", "evidence": evidence}
        if unsat and non100:
            note = "（" + "、".join(f"{n}不满意" for n in unsat) + \
                   "｜非100：" + "、".join(f"{n}({s}分)" for n, s in non100) + "）"
            return {"result": "不合格" + note, "qualified": False, "score": "",
                    "confidence": "high", "evidence": evidence}
        return {"result": "合格", "qualified": True, "score": "",
                "confidence": "high", "evidence": evidence}

    # 兜底（无分项结构）：原整体判定
    has_fail = bool(re.search(r"不通过|不合格|不及格|不满意|未通过|不达标", txt))
    has_pass = bool(re.search(r"通过|合格|满意|及格", txt))
    score = ""
    m = re.search(r"成绩\s*[:：]?\s*(\d{1,3})\s*[％%]?", txt)
    if m:
        score = m.group(1)
    if not score:
        m2 = re.search(r"得分\s*[:：]?\s*(\d{1,3})", txt)
        if m2:
            score = m2.group(1)
    if has_fail and not has_pass:
        return {"result": "不合格", "qualified": False, "score": score, "confidence": "high", "evidence": evidence}
    if has_fail and has_pass:
        return {"result": "含不合格项(待复核)", "qualified": False, "score": score, "confidence": "medium", "evidence": evidence}
    if has_pass and not has_fail:
        result = f"成绩{score}%" if score else "合格"
        return {"result": result, "qualified": True, "score": score, "confidence": "high", "evidence": evidence}
    return {"result": "", "qualified": None, "score": "", "confidence": "low", "evidence": evidence}


@router.post("/report/{plan_id}")
async def upload_eqa_report(
    plan_id: int,
    file: UploadFile = File(...),
    score: str = Form(""),
    qualified: bool = Form(False),
    result: str = Form(""),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导入某条质评计划的 PDF 报告（卫健委/北京市），保存到 data/eqa_reports/ 并入库；
    同时回填成绩/得分与是否合格。"""
    plan = db.query(EqaPlan).filter(EqaPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")
    fname = (file.filename or "").lower()
    if not fname.endswith(".pdf"):
        raise HTTPException(status_code=400, detail="仅支持 PDF 质评报告")
    os.makedirs(EQA_REPORT_DIR, exist_ok=True)
    safe = f"{plan_id}_{int(datetime.utcnow().timestamp())}.pdf"
    dest = EQA_REPORT_DIR / safe
    content = await file.read()
    with open(dest, "wb") as f:
        f.write(content)
    # 解析口子：抽取成绩/合格/得分（失败不影响导入）；传入 item 以按子项识别不适用
    parsed = {}
    try:
        parsed = parse_eqa_report(str(dest), item_text=plan.item)
    except Exception:
        parsed = {}
    # 删除该计划旧报告文件
    if plan.report_file:
        old = EQA_REPORT_DIR / os.path.basename(plan.report_file)
        if old.exists():
            try:
                old.unlink()
            except Exception:
                pass
    plan.report_file = f"eqa_reports/{safe}"
    # 回填优先级：表单显式输入 > 自动解析；空白字段用解析结果补全
    if score != "":
        plan.score = score
    elif parsed.get("score"):
        plan.score = parsed["score"]
    if result != "":
        plan.result = result
    elif parsed.get("result"):
        plan.result = parsed["result"]
    if qualified:
        plan.qualified = True
    elif parsed.get("qualified") is not None:
        plan.qualified = parsed["qualified"]
    db.commit()
    db.refresh(plan)
    return EqaPlanRead.model_validate(plan)


@router.get("/reports-merge")
def merge_eqa_reports(
    year: int,
    half: int = 0,
    category: str = "生化+凝血",
    department: str = "卫健委",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """结果报告合并打印：将某（年度 × 半年 × 质评部门 × 专业组）下所有已导入报告的
    质评计划 PDF 合并为一份，便于集中打印。返回 inline PDF（浏览器可直接打印）。

    - half：0=全年 / 1=上半年（1-6月，按 due_date 月份）/ 2=下半年（7-12月）
    - category：生化+凝血 / 免疫（与总结统计同一分类口径）
    - department：卫健委 / 北京市（与总结统计同一部门口径）
    """
    from urllib.parse import quote

    cat_target = _normalize_category(category)
    dept_target = _normalize_department(department)

    q = db.query(EqaPlan).filter(EqaPlan.year == year)
    if half == 1:
        q = q.filter(func.substr(EqaPlan.due_date, 6, 2).in_(_H1))
    elif half == 2:
        q = q.filter(func.substr(EqaPlan.due_date, 6, 2).in_(_H2))
    if dept_target:
        q = q.filter(EqaPlan.org == dept_target)
    plans = q.order_by(EqaPlan.program, EqaPlan.round_no, EqaPlan.due_date).all()
    # 只保留目标分类 + 已导入报告且磁盘文件存在者
    files = []
    for p in plans:
        if _category_of(p.group) != cat_target or not p.report_file:
            continue
        path = EQA_REPORT_DIR / os.path.basename(p.report_file)
        if path.exists():
            files.append((p, path))
    if not files:
        raise HTTPException(status_code=404, detail="该范围下没有可合并的已导入报告")

    from pypdf import PdfReader, PdfWriter

    writer = PdfWriter()
    for p, path in files:
        try:
            reader = PdfReader(str(path))
            for page in reader.pages:
                writer.add_page(page)
        except Exception:
            # 跳过损坏/非 PDF 文件，保证其余报告仍可合并
            continue
    out = BytesIO()
    writer.write(out)
    out.seek(0)

    slug = f"{year}_{_ORG_SLUG.get(dept_target, dept_target)}_{_CATEGORY_SLUG.get(cat_target, cat_target)}_结果报告合并"
    fname = f"{slug}.pdf"
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename*=UTF-8''{quote(fname)}"},
    )


@router.get("/report/{plan_id}")
def download_eqa_report(
    plan_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """下载某条质评计划已导入的 PDF 报告。"""
    plan = db.query(EqaPlan).filter(EqaPlan.id == plan_id).first()
    if not plan or not plan.report_file:
        raise HTTPException(status_code=404, detail="未导入报告")
    path = EQA_REPORT_DIR / os.path.basename(plan.report_file)
    if not path.exists():
        raise HTTPException(status_code=404, detail="报告文件缺失")
    return FileResponse(
        str(path),
        filename=os.path.basename(plan.report_file),
        media_type="application/pdf",
    )


@router.delete("/report/{plan_id}")
def delete_eqa_report(
    plan_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """删除某条质评计划已导入的 PDF 报告。"""
    plan = db.query(EqaPlan).filter(EqaPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")
    if plan.report_file:
        path = EQA_REPORT_DIR / os.path.basename(plan.report_file)
        if path.exists():
            try:
                path.unlink()
            except Exception:
                pass
        plan.report_file = ""
        db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# 导出年度计划清单（Excel）
# ---------------------------------------------------------------------------
@router.get("/export")
def export_eqa(
    year: int | None = None,
    org: str | None = None,
    group: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from urllib.parse import quote

    q = db.query(EqaPlan)
    if year is not None:
        q = q.filter(EqaPlan.year == year)
    if org:
        q = q.filter(EqaPlan.org == org)
    if group:
        q = q.filter(EqaPlan.group == group)
    plans = q.order_by(EqaPlan.year.desc(), EqaPlan.due_date, EqaPlan.org, EqaPlan.program).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "室间质评计划"
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    headers = ["年度", "组织机构", "专业组", "项目组", "细项", "轮次", "样本检测日期",
               "上报截止日期", "是否上报", "成绩", "是否合格", "报告", "备注"]
    report_map = {p.id: ("已导入" if p.report_file else "") for p in plans}
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    for i, p in enumerate(plans, 2):
        vals = [
            p.year, p.org, p.group, p.program, p.item, p.round_no, p.sample_date,
            p.due_date, "是" if p.returned else "否", p.result or p.score,
            "合格" if p.qualified else ("不合格" if (p.result or p.score) else ""),
            report_map.get(p.id, ""), p.note,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.alignment = center
            cell.border = border
    widths = [8, 22, 10, 16, 18, 10, 16, 16, 10, 12, 10, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    fname = f"室间质评计划_{year if year else 'ALL'}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"},
    )


# ---------------------------------------------------------------------------
# 一键复制上一年全部计划到目标年（批量，避免逐条 after_write 的 O(n^2) 通知重建）
# ---------------------------------------------------------------------------
def _shift_year(d, delta):
    """把 YYYY-MM-DD 类日期的年份整体偏移 delta 年（如 2025-03-19 -> 2026-03-19）。
    非日期字符串或空值原样返回。"""
    if not d:
        return d
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", str(d))
    if not m:
        return d
    return f"{int(m.group(1)) + delta:04d}-{m.group(2)}-{m.group(3)}"


@router.post("/copy-prev-year")
def copy_prev_year_eqa(
    target_year: int | None = None,
    source_year: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """一键把 source_year（默认 target_year-1）的全部质评计划复制到目标年（默认今年）。

    - 新计划 year 改为 target_year；回报/报告字段（returned/result/qualified/score/report_file）清空。
    - 按 (org, program, item, round_no) 去重：目标年已存在相同 key 的计划自动跳过（避免重复点击产生重复数据）。
    - 批量插入后仅 refresh 一次上报提醒（性能考量）。
    """
    if target_year is None:
        target_year = datetime.now().year
    if source_year is None:
        source_year = target_year - 1
    src = db.query(EqaPlan).filter(EqaPlan.year == source_year).all()
    if not src:
        return {
            "copied": 0,
            "skipped": 0,
            "target_year": target_year,
            "source_year": source_year,
            "message": f"上一年（{source_year}）无计划可复制",
        }
    existing = set(
        db.query(EqaPlan.org, EqaPlan.program, EqaPlan.item, EqaPlan.round_no)
        .filter(EqaPlan.year == target_year)
        .all()
    )
    new_objs = []
    skipped = 0
    for p in src:
        key = (p.org, p.program, p.item, p.round_no)
        if key in existing:
            skipped += 1
            continue
        new_objs.append(
            EqaPlan(
                year=target_year,
                org=p.org,
                program=p.program,
                group=p.group,
                item=p.item,
                round_no=p.round_no,
                sample_date=_shift_year(p.sample_date, target_year - source_year),
                due_date=_shift_year(p.due_date, target_year - source_year),
                note=p.note,
                returned=False,
                result="",
                qualified=False,
                score="",
                report_file="",
            )
        )
    if new_objs:
        db.add_all(new_objs)
        db.commit()
    # 末尾统一刷新一次提醒（非逐条），避免 O(n^2)
    refresh_eqa_notifications(db)
    return {
        "copied": len(new_objs),
        "skipped": skipped,
        "target_year": target_year,
        "source_year": source_year,
    }


# ---------------------------------------------------------------------------
# 逐项「录入结果」：样本×项目 矩阵表格（类似卫健委/北京市下发的成绩回报单）
# 以 JSON 存于 eqa_plans.result_data：{samples:[], items:[], cells:{样本:{项目:值}},
# meta:{tester,reviewer,test_date,return_date,note}}
# ---------------------------------------------------------------------------
class EqaResultData(BaseModel):
    samples: list[str] = []
    items: list[str] = []
    units: dict[str, str] = {}  # 项目名 -> 单位（上报单位，换算项目为 to）
    conv: dict[str, dict] = {}  # 项目名 -> {factor, from, to}，需单位换算后上报的项目
    cells: dict[str, dict[str, str]] = {}  # 原值（仪器单位）
    cells_report: dict[str, dict[str, str]] = {}  # 换算后的上报值
    meta: dict = {}


def _parse_round(round_no: str) -> int:
    """从轮次文本提取轮次数字：'第1次'/'2026-1'/'2026年第2次' -> 1/1/2；失败回 1。"""
    if not round_no:
        return 1
    nums = re.findall(r"\d+", round_no)
    return int(nums[-1]) if nums else 1


def _norm_item(s: str) -> str:
    """项目名归一化：去全/半角括号差异与空白，转小写，便于与项目查询库匹配。"""
    s = (s or "").strip()
    s = s.replace("（", "(").replace("）", ")").replace("　", " ").replace(" ", "")
    return s.lower()


def _lookup_units(names, db: Session) -> dict:
    """按项目名/别名匹配项目查询库(test_items)，返回 {项目名: 单位}。

    匹配策略：先精确（归一化后相等），再子串包含。覆盖率约 74%，
    未命中返回 ''，由前端允许手动填写单位。
    """
    if not names:
        return {}
    rows = db.query(TestItem.name, TestItem.aliases, TestItem.unit).all()
    index = []
    for name, aliases, unit in rows:
        keys = {_norm_item(name)}
        for a in (aliases or "").split(","):
            a = a.strip()
            if a:
                keys.add(_norm_item(a))
        index.append((keys, unit or ""))
    out = {}
    for n in names:
        nn = _norm_item(n)
        unit = ""
        if nn:
            for keys, u in index:
                if nn in keys:
                    unit = u
                    break
            if not unit:
                for keys, u in index:
                    for k in keys:
                        if k and len(k) >= 2 and len(nn) >= 2 and (k in nn or nn in k):
                            unit = u
                            break
                    if unit:
                        break
        out[n] = unit
    return out


# ---------------------------------------------------------------------------
# EQA 上报单位换算表（部分激素项目仪器测原单位，需换算后上报）
# factor：原值 × factor = 上报值；from：仪器单位（提示用）；to：上报单位。
# names：归一化别名集合（含项目库/计划里可能出现的写法，如 uE3/e2）。
# ---------------------------------------------------------------------------
EQA_UNIT_CONVERSIONS = [
    {"names": ["游离雌三醇", "雌三醇", "ue3"], "factor": 3.467, "from": "pg/mL", "to": "nmol/L"},
    {"names": ["雌二醇", "e2"], "factor": 3.67, "from": "pg/mL", "to": "pmol/L"},
    {"names": ["孕酮", "黄体酮", "p"], "factor": 3.18, "from": "ng/mL", "to": "nmol/L"},
    {"names": ["pth", "甲状旁腺激素"], "factor": 0.106, "from": "pg/mL", "to": "pmol/L"},
    {"names": ["皮质醇", "cortisol"], "factor": 27.64, "from": "μg/dL", "to": "nmol/L"},
    {"names": ["acth"], "factor": 0.2202, "from": "pg/mL", "to": "pmol/L"},
    {"names": ["醛固酮", "aldosterone", "ald"], "factor": 2.775, "from": "pg/mL", "to": "nmol/L"},
    {"names": ["胰岛素", "insulin"], "factor": 6.965, "from": "μU/mL", "to": "pmol/L"},
    {"names": ["c肽", "c-peptide", "连接肽"], "factor": 0.333, "from": "ng/mL", "to": "nmol/L"},
    {"names": ["vd", "25-ohvd", "维生素d", "25羟维生素d"], "factor": 2.494, "from": "ng/mL", "to": "nmol/L"},
    {"names": ["肾素", "renin"], "factor": 1.2, "from": "ng/mL", "to": "uIU/mL"},
]


def _match_conversion(name: str):
    """按项目名/别名匹配是否需要单位换算，返回 {factor, from, to} 或 None。"""
    nn = _norm_item(name)
    if not nn:
        return None
    for rule in EQA_UNIT_CONVERSIONS:
        if nn in rule["names"]:
            return {"factor": rule["factor"], "from": rule["from"], "to": rule["to"]}
        for alias in rule["names"]:
            if len(alias) >= 2 and alias in nn:
                return {"factor": rule["factor"], "from": rule["from"], "to": rule["to"]}
    return None


# ---------------------------------------------------------------------------
# EQA 定性试验（COI）阴阳性判定
# 覆盖：肝炎标志物、感染性疾病血清学标志物系列A/B/C（感A/感B/感C）等。
# 单位不改，仅按 COI（S/CO）判定阴阳性：
#   - 常规项目：COI > 1 → 阳性(P)，否则 阴性(N)
#   - 例外（抑制/竞争法抗体，COI 越低越阳）：HBeAb、HBcAb → COI < 1 → 阳性(P)，否则 阴性(N)
# result_data 中保留 COI 原值（存于 cells），并新增 qual_pn = {样本:{项目:'P'/'N'}}。
# ---------------------------------------------------------------------------
EQA_QUALITATIVE_PROGRAM_KEYWORDS = ["肝炎", "感染"]  # program 含这些字 → 定性试验（COI）
EQA_QUALITATIVE_REVERSE_ITEMS = [
    "hbeab", "抗-hbe", "抗hbe", "hbe抗体", "e抗体", "乙肝e抗体",
    "hbcab", "抗-hbc", "抗hbc", "hbc抗体", "核心抗体", "乙肝核心抗体",
]


def _is_qualitative_program(program: str) -> bool:
    """program 名含 肝炎 / 感染 → 定性试验（按 COI 判阴阳）。"""
    p = (program or "").lower()
    return any(k in p for k in EQA_QUALITATIVE_PROGRAM_KEYWORDS)


def _is_reverse_item(item: str) -> bool:
    """HBeAb / HBcAb 等（COI<1 判阳）项目。注意与 HBeAg/HBcAg 区分。"""
    nn = _norm_item(item)
    if not nn:
        return False
    for k in EQA_QUALITATIVE_REVERSE_ITEMS:
        kk = _norm_item(k)
        if kk == nn or kk in nn or nn in kk:
            return True
    return False


def _qualitative_pn(value, reverse: bool) -> str:
    """返回 'P'(阳性) / 'N'(阴性) / ''（无值/非法）。

    value 兼容多种历史写法：可带 '<'/'≤'（低于 cutoff）、可带末尾 '/P' '/N'
    （旧数据把 数值/阴阳 写在一起，如 '908.2/P'、'<2/N'、'0.016/P'），
    解析时自动忽略后缀、识别低于 cutoff。
    """
    if value is None:
        return ""
    s = str(value).strip()
    # 兼容旧数据：去掉末尾 /P /N
    s = re.sub(r"/[PN]$", "", s, flags=re.IGNORECASE).strip()
    if not s:
        return ""
    below = bool(re.match(r"^(<|≤|<=|小于)", s))
    s2 = re.sub(r"^(<|≤|<=|小于)", "", s).strip()
    try:
        v = float(s2)
    except (TypeError, ValueError):
        return ""
    if below:
        # 低于 cutoff：常规项目为阴(N)，反向项目(HBeAb/HBcAb)为阳(P)
        return "N" if not reverse else "P"
    if reverse:
        return "P" if v < 1 else "N"
    return "P" if v > 1 else "N"


def _compute_qual_pn(payload: dict) -> dict:
    """对定性试验计划，按 cells（COI 原值）计算 qual_pn 阴阳性表。"""
    items = payload.get("items") or []
    cells = payload.get("cells") or {}
    pn = {}
    for s, row in cells.items():
        pn[s] = {}
        for it in items:
            val = (row or {}).get(it, "")
            pn[s][it] = _qualitative_pn(val, _is_reverse_item(it))
    return pn


def _split_items(raw: str) -> list:
    """把 plan.item 拆成具体项目列表，并清理 '具体项目：/项目：/检测项目：' 前缀。"""
    raw = re.sub(r"^(具体项目|项目|检测项目)[：:]\s*", "", raw or "")
    items = []
    for part in re.split(r"[、,，/]", raw):
        part = part.strip()
        if part:
            items.append(part)
    return items


# ---- 标本感知辅助（总结细项去重用）----
def _specimen_of_program(program: str) -> str:
    """按项目组名推断标本类型（用于总结细项按标本分开计数，而非同名合并）。"""
    p = program or ""
    if "尿液" in p:
        return "尿液"
    if "脑脊液" in p:
        return "脑脊液"
    if "血气" in p:
        return "血气"
    if any(k in p for k in ("凝血", "D-二聚体", "纤维蛋白（原）降", "抗凝蛋白",
                             "抗凝血因子Xa", "血管性血友病")):
        return "血浆"
    return "血清"


_SPEC_RE = re.compile(r"[（(](尿液|脑脊液|胸腹水|胸水|腹水|血浆|血清)[)）]")


def _split_specimen(token: str, spec_plan: str):
    """单个细项 token → (标本, 分析物名)。token 自带标本后缀时优先，否则用所属项目组标本。"""
    m = _SPEC_RE.search(token or "")
    if m:
        spec = m.group(1)
        an = _SPEC_RE.sub("", token).strip()
    else:
        spec = spec_plan
        an = (token or "").strip()
    return spec, an


def _build_skeleton(plan: EqaPlan, db: Session = None) -> dict:
    """无已存数据时，按计划信息生成可填骨架。

    样本编号规则：年份(4位) + 轮次(1位) + 样本序号(1位)。
    例：2026 年第 1 轮 5 个样本 -> 202611~202615；2025 年第 1 轮 -> 202511~202515。
    项目来自 plan.item（已清理前缀），单位来自项目查询库匹配。
    """
    items = _split_items(plan.item)
    units = _lookup_units(items, db) if db is not None else {}
    conv = {}
    for it in items:
        c = _match_conversion(it)
        if c:
            conv[it] = c
            units[it] = c["to"]  # 上报单位覆盖为换算目标单位
    year = plan.year or datetime.now().year
    rnd = _parse_round(plan.round_no)
    samples = [f"{year}{rnd}{i}" for i in range(1, 6)]  # 默认 5 个样本
    return {
        "samples": samples,
        "items": items,
        "units": units,
        "conv": conv,
        "cells": {},
        "cells_report": {},
        "meta": {
            "tester": "",
            "reviewer": "",
            "test_date": plan.sample_date or "",
            "return_date": plan.due_date or "",
            "note": "",
        },
    }


@router.get("/{plan_id}/result")
def get_eqa_result(
    plan_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """读取某计划的逐项结果矩阵；若无已存数据返回基于计划的骨架。"""
    plan = db.query(EqaPlan).filter(EqaPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")
    rd = _build_skeleton(plan, db)
    if plan.result_data:
        try:
            stored = json.loads(plan.result_data)
            if isinstance(stored, dict):
                rd = stored
        except Exception:
            pass
    # 自愈：已存数据项目列表为空、但计划本身有细项时，按当前计划重建骨架，
    # 避免「录入结果」打开后项目为空（历史空数据自愈 + 对未来类似情况免疫）。
    if isinstance(rd, dict) and (not rd.get("items")) and plan.item and plan.item.strip():
        rd = _build_skeleton(plan, db)
    # 已存数据若缺 units，则按当前项目查询库补算，保证单位能显示
    if isinstance(rd, dict) and not rd.get("units") and rd.get("items"):
        rd["units"] = _lookup_units(rd["items"], db)
    # 回填单位换算信息（旧数据 / 新数据都保证 conv 与上报单位存在）
    if isinstance(rd, dict) and rd.get("items"):
        rd.setdefault("conv", {})
        rd.setdefault("cells_report", {})
        for it in rd["items"]:
            c = _match_conversion(it)
            if c and it not in rd["conv"]:
                rd["conv"][it] = c
                rd["units"][it] = c["to"]
    return {
        "result_data": rd,
        "plan": {
            "id": plan.id,
            "org": plan.org,
            "program": plan.program,
            "item": plan.item,
            "round_no": plan.round_no,
            "year": plan.year,
            "sample_date": plan.sample_date,
            "due_date": plan.due_date,
        },
    }


@router.put("/{plan_id}/result")
def put_eqa_result(
    plan_id: int,
    body: EqaResultData,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """保存某计划的逐项结果矩阵（覆盖写入 JSON）。

    定性试验（肝炎/感染血清学）额外按 COI 计算 qual_pn（阴阳性 P/N）：
    单位不变，仅按 COI 值判阴阳；HBeAb/HBcAb 取反向规则（COI<1 为阳）。
    """
    plan = db.query(EqaPlan).filter(EqaPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(status_code=404, detail="计划不存在")
    payload = body.model_dump()
    # 定性试验：保留 COI 原值（cells）并自动生成阴阳性表 qual_pn
    if _is_qualitative_program(plan.program):
        payload["qualitative"] = True
        payload["qual_pn"] = _compute_qual_pn(payload)
    else:
        payload.pop("qualitative", None)
        payload.pop("qual_pn", None)
    plan.result_data = json.dumps(payload, ensure_ascii=False)
    db.commit()
    return {"ok": True, "result_data": payload}


# ---------------------------------------------------------------------------
# 路由重排（必须放在所有路由定义之后）：把不含路径参数（{...}）的静态路由
# （/alerts、/summary、/summary-text、/export）排到最前，避免被 make_router
# 生成的 GET /{item_id} 等参数路由吞掉（否则 /summary 被当成 int 解析 → 422）。
# sort 稳定，保持组内相对顺序。
# ---------------------------------------------------------------------------
router.routes.sort(key=lambda r: ("{" in getattr(r, "path", ""),))
