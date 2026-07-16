from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import or_
from io import BytesIO
import io
import openpyxl
from collections import Counter

from ...core.crud_base import make_router
from ...models.test_item import TestItem
from ...models.user import User
from ...schemas import TestItemCreate, TestItemRead, TestItemUpdate
from ...core.database import get_db
from ...core.security import get_current_user
from ...core.brand import extract_brand, resolve_brand

router = make_router(
    TestItem,
    TestItemRead,
    TestItemCreate,
    TestItemUpdate,
    search_fields=["code", "name", "aliases", "category", "method", "instrument", "instrument_group"],
    filter_fields=["category", "brand", "specimen", "method", "instrument"],
    prefix="/test-items",
    write_roles=("admin",),
)

# 中文表头 -> 英文字段名（与导出模板一致，忽略表头尾部 * 号）
_CN2EN = {
    "项目编号": "code", "项目名称": "name", "别名": "aliases", "类别": "category",
    "标本类型": "specimen", "方法学": "method", "单位": "unit", "参考范围": "reference",
    "收费": "fee", "使用仪器": "instrument", "仪器组": "instrument_group",
    "线性范围": "linear_range", "稀释倍数": "dilution_fold", "可报告范围": "reportable_range",
    "稀释液": "diluent", "校准品": "calibrator", "溯源性": "traceability", "品牌": "brand",
    "最近更新": "last_update", "溶血干扰": "interference_hemolysis",
    "胆红素干扰": "interference_bilirubin", "脂血干扰": "interference_lipemia",
}


@router.post("/import")
async def import_test_items(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """批量导入检验项目：按项目名称匹配，已存在则更新、不存在则新增。"""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")
    try:
        raw = await file.read()
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
        ws = wb.active
        grid = list(ws.iter_rows(values_only=True))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"读取文件失败：{e}")
    if not grid:
        raise HTTPException(status_code=400, detail="文件为空")

    header = [str(c).strip().rstrip("*").strip() if c is not None else "" for c in grid[0]]
    en_header = [_CN2EN.get(h, "") for h in header]

    created = updated = skipped = 0
    errors = []
    for row in grid[1:]:
        if row is None:
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        payload = {}
        for idx, en in enumerate(en_header):
            if not en or idx >= len(row):
                continue
            val = row[idx]
            payload[en] = "" if val is None else str(val).strip()
        name = payload.get("name", "")
        if not name:
            skipped += 1
            errors.append("存在缺项目名称的行，已跳过")
            continue
        existing = db.query(TestItem).filter(TestItem.name == name).first()
        if existing:
            changed = False
            for k, v in payload.items():
                if v != "" and getattr(existing, k) != v:
                    setattr(existing, k, v)
                    changed = True
            if changed:
                updated += 1
            else:
                skipped += 1
        else:
            db.add(TestItem(**payload))
            created += 1
        db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors[:20]}


@router.get("/stats")
async def test_items_stats(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """统计分类：按类别、按品牌、按仪器的计数，供前端概览展示。"""
    rows = db.query(TestItem).all()
    cat_counter = Counter(r.category or "未分类" for r in rows)
    brand_counter = Counter(resolve_brand(r.calibrator, r.brand) or "未标识" for r in rows)
    def _norm_inst(r):
        v = (r.instrument or "").strip()
        return "未设置" if v in ("", "/") else v
    inst_counter = Counter(_norm_inst(r) for r in rows)
    return {
        "total": len(rows),
        "category_counts": dict(cat_counter.most_common()),
        "brand_counts": dict(brand_counter.most_common()),
        "instrument_counts": dict(inst_counter.most_common()),
    }


@router.get("/export")
async def export_test_items(
    q: str = "",
    category: str = "",
    brand: str = "",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导出项目汇总清单（Excel）：支持按关键词/类别/品牌筛选，按类别排序，并附分类统计。"""
    query = db.query(TestItem)
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                TestItem.name.ilike(like),
                TestItem.code.ilike(like),
                TestItem.aliases.ilike(like),
                TestItem.method.ilike(like),
                TestItem.instrument.ilike(like),
                TestItem.instrument_group.ilike(like),
            )
        )
    if category:
        query = query.filter(TestItem.category == category)
    if brand:
        query = query.filter(or_(TestItem.brand == brand, TestItem.calibrator.ilike(f"%{brand}%")))
    rows = query.order_by(TestItem.category, TestItem.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目汇总清单"
    headers = [
        "项目编号", "项目名称", "品牌", "类别", "标本类型", "方法学", "单位",
        "参考范围", "线性范围", "可报告范围", "校准品", "溯源性", "使用仪器", "仪器组",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.code, r.name, resolve_brand(r.calibrator, r.brand), r.category, r.specimen,
            r.method, r.unit, r.reference, r.linear_range, r.reportable_range,
            r.calibrator, r.traceability, r.instrument, r.instrument_group,
        ])

    # 分类统计 sheet
    stat = wb.create_sheet("分类统计")
    cat_counter = Counter(r.category or "未分类" for r in rows)
    brand_counter = Counter(resolve_brand(r.calibrator, r.brand) or "未标识" for r in rows)
    stat.append(["一、按类别统计", ""])
    stat.append(["类别", "数量"])
    for k, v in cat_counter.most_common():
        stat.append([k, v])
    stat.append([])
    stat.append(["二、按品牌统计", ""])
    stat.append(["品牌", "数量"])
    for k, v in brand_counter.most_common():
        stat.append([k, v])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=test_items_summary.xlsx",
            "Cache-Control": "no-store",
        },
    )
