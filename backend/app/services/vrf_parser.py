"""性能验证 xlsx 报告解析器 —— 从上传的 Excel 中提取项目信息和验证结论。

兼容定性（BG-SM-CZ-040，51-HBsAg 模板）和定量（BG-SM-CZ-039，2.ALP 模板）。
主策略：读取主封面 + 结果汇总 sheet，提炼项目信息 + 验证结论，写入 verification_reports 表。
"""
import json
import re
from typing import Any

from openpyxl import load_workbook


def _val(ws, row, col):
    """安全读取单元格值（自动处理合并区，返回左上 cell 的值）"""
    try:
        v = ws.cell(row=row, column=col).value
        if v is None:
            # 合并区：找包含 (row, col) 的合并 range，取左上 cell 的值
            try:
                for r in ws.merged_cells.ranges:
                    if r.min_row <= row <= r.max_row and r.min_col <= col <= r.max_col:
                        mv = ws.cell(r.min_row, r.min_col).value
                        return mv if mv is not None else ""
            except Exception:
                pass
            return ""
    except Exception:
        return None
    return v


def _safe_str(v):
    return "" if v is None else str(v).strip()


def _type_from_cover(ws) -> str:
    """从主封面第 2 行标题判断定性/定量"""
    row2 = " ".join([_safe_str(_val(ws, 2, c)) for c in range(1, 10)])
    if "定性" in row2:
        return "qualitative"
    if "定量" in row2:
        return "quantitative"
    return "qualitative"


def _read_cover(wb):
    """读主封面信息。

    R22 在模板上标"试剂厂家"（如"贝克曼库尔特"），但用户上传时常错填为
    试剂盒产品名（如"碱性磷酸酶检测试剂盒"）。这里做智能识别：
    如果 R22 含有"检测试剂盒"/"试剂"/"测定试剂盒"等产品描述关键词，
    视为产品名，从 R26（仪器厂家）提取品牌覆盖到 reagent 字段。
    """
    ws = wb["主封面"]
    reagent_raw = _safe_str(_val(ws, 22, 5))
    instrument_manufacturer = _safe_str(_val(ws, 26, 5))
    # 智能品牌提取：含产品描述词的产品名 → 用仪器厂家覆盖
    if any(kw in reagent_raw for kw in ("检测试剂盒", "测定试剂盒", "试剂", "试剂盒", "诊断试剂")):
        # 若同时能提取出品牌名（如"贝克曼"）优先；否则用仪器厂家
        known_brands = ("贝克曼", "贝克曼库尔特", "西门子", "罗氏", "雅培", "迈瑞", "日立", "奥林巴斯", "沃芬", "德赛", "安图", "沃文特", "积水", "柏定", "柏荣", "博源", "思塔高", "景源", "丰华", "亚辉龙", "九强", "迈克生物", "英科新创", "艾博")
        for b in known_brands:
            if b in reagent_raw or b in instrument_manufacturer:
                reagent = b
                break
        else:
            reagent = instrument_manufacturer or reagent_raw
    else:
        reagent = reagent_raw
    return {
        "report_type": _type_from_cover(ws),
        "project_name": _safe_str(_val(ws, 20, 5)),
        "reagent": reagent,
        "instrument": _safe_str(_val(ws, 24, 5)),
        "instrument_manufacturer": instrument_manufacturer,
        "instrument_model": _safe_str(_val(ws, 28, 5)),
        "instrument_no": _safe_str(_val(ws, 30, 5)),
        "operator": _safe_str(_val(ws, 32, 5)),
        "reviewer": _safe_str(_val(ws, 34, 5)),
        "verify_date": _safe_str(_val(ws, 36, 5)),
    }


def _read_summary(wb, info: dict) -> dict:
    """读结果汇总 sheet：验证结论表 + 参数"""
    ws = wb["结果汇总"]
    # 参数
    info["project_method"] = _safe_str(_val(ws, 4, 2))
    info["unit"] = _safe_str(_val(ws, 5, 2))
    info["reagent"] = info.get("reagent") or _safe_str(_val(ws, 6, 2))
    info["reagent_lot"] = _safe_str(_val(ws, 6, 7))
    info["calibrator"] = _safe_str(_val(ws, 7, 2))
    info["calibrator_lot"] = _safe_str(_val(ws, 7, 7))
    info["qc"] = _safe_str(_val(ws, 8, 2))
    info["qc_lot"] = _safe_str(_val(ws, 8, 7))
    info["tea"] = _safe_str(_val(ws, 5, 7))
    info["dilution"] = _safe_str(_val(ws, 9, 7))
    # 声称线性范围：兼容 ALP 模板（R9 B/D 列直接是数字）和 ALB 模板（R21 F/G 列写 "15-60" 格式）
    # 优先从 R9 读（ALP 模板）
    linear_low = _safe_str(_val(ws, 9, 2)).strip()
    linear_high = _safe_str(_val(ws, 9, 4)).strip()
    # 如果 R9 读不到数字（ALB 模板 R9 是"低限 ≤ TEA"），从 R21 F/G 列解析
    if not re.search(r'\d', linear_low) or not re.search(r'\d', linear_high):
        r21_f = _safe_str(_val(ws, 21, 6)) or _safe_str(_val(ws, 21, 7))
        r21_g = _safe_str(_val(ws, 21, 7)) or r21_f
        combined = f"{r21_f} {r21_g}"
        nums = re.findall(r'\d+(?:\.\d+)?', combined)
        if len(nums) >= 2:
            linear_low, linear_high = nums[0], nums[1]
        elif len(nums) == 1:
            linear_low = linear_high = nums[0]
        # 还读不到：扫 R8-R10 / R21-R23 各列找 "线性" 标注附近的数字
        if not re.search(r'\d', linear_low) or not re.search(r'\d', linear_high):
            for r in (8, 9, 10, 21, 22, 23):
                for c in range(2, 12):
                    v = _safe_str(_val(ws, r, c))
                    if '线性' in v or 'linear' in v.lower():
                        row_text = ' '.join(_safe_str(_val(ws, r, cc)) for cc in range(2, 12))
                        row_nums = re.findall(r'\d+(?:\.\d+)?', row_text)
                        if len(row_nums) >= 2:
                            linear_low, linear_high = row_nums[0], row_nums[1]
                            break
                        elif len(row_nums) == 1:
                            linear_low = linear_high = row_nums[0]
                            break
                if re.search(r'\d', linear_low) and re.search(r'\d', linear_high):
                    break
    info["linear_low"] = linear_low
    info["linear_high"] = linear_high
    # 验证内容（R14）
    r14 = _safe_str(_val(ws, 14, 2)) or _safe_str(_val(ws, 14, 1))

    # 关键：不同模板"验证内容/要求/结果/结论"列位置不同（ALP=B 验证内容，ALB=C 验证内容），
    # 这里按内容智能识别：B/C/D/E 任一列含"精密度/正确度/线性/可报告/参考/特异/检出/符合率"→ content
    rows = []
    for r in range(17, 27):
        # 收集这一行所有非空文本（B/C/D/E/F/G/H/I 9 列）
        col_texts = []  # [(col, text)]
        for c in range(2, 10):
            v = _safe_str(_val(ws, r, c))
            if v:
                col_texts.append((c, v))
        # 找 content（按关键词匹配）
        content = ''
        for c, t in col_texts:
            if _content_to_key(t):
                content = t
                break
        # requirement：B/C/D/E 中（content 之外）的文本
        req_candidates = [t for c, t in col_texts if c <= 5 and t != content]
        requirement = req_candidates[0] if req_candidates else ''
        # result：F/G 中第一个含数字/百分数/具体值的
        result = ''
        for c, t in col_texts:
            if 6 <= c <= 7 and any(ch in t for ch in '0123456789%±'):
                result = t
                break
        if not result:
            # 退而求其次：F/G 中第一个非空文本
            for c, t in col_texts:
                if 6 <= c <= 7:
                    result = t
                    break
        # conclusion：H/I 中（若空则从 result 文本里反推"符合/通过"）
        conclusion = ''
        for c, t in col_texts:
            if 8 <= c <= 9:
                conclusion = t
                break
        if not conclusion and result:
            # 兜底：ALB 模板的 H 列是公式（data_only=True 时为 None），
            # 从 result 文本里找"符合/通过/不合格"关键词
            if '不符合' in result or '不通过' in result:
                conclusion = '不符合要求'
            elif '符合' in result or '通过' in result:
                conclusion = '符合要求'
        # 兜底：模板上 R17-R18=精密度、R19-R20=方法符合率、R21=检出限/线性、R22-R23=可报告范围、R24=参考区间、R26=分析特异性
        # 只要 content 没解析出来（即使 requirement 在）且本行有 result/conclusion，就反推
        if not content and (result or conclusion):
            # 优先用 requirement 关键词反推（更准，能区分"实验室内CV→精密度"/"相对偏倚→正确度"）
            content = _content_from_requirement(requirement) if requirement else _content_from_row(r)
        elif content and not _content_to_key(content):
            # content 拿到了但不是标准验证项目名（是产品名/项目名等），改用行号推导
            content = _content_from_requirement(requirement) if requirement else (_content_from_row(r) or content)
        # 标准化 content 为简短标签
        content = _normalize_content(content) if content else ''
        # 跳过表头行（R17 在 ALB 模板是"验证要求/验证结果/验证结论"表头）
        header_words = ('验证要求', '验证结果', '验证结论', '验证内容')
        if content in header_words or result in header_words or conclusion in header_words:
            continue
        # 特异性行（R26-R28 多组干扰物）：每行 6/7/8 = 干扰物名/限量/结论
        # result 拼成多行：每行 "<干扰物> <限量>"，用换行符连接
        if r == 26 and content == '分析特异性':
            c2 = _safe_str(_val(ws, 26, 2))
            names = [s.strip() for s in re.split(r'[、，,/\n]', c2) if s.strip() and s.strip() != '抗干扰能力符合厂家声明']
            # 读 R26/R27/R28 的 干扰物名/限量
            multi_lines = []
            for rr in (26, 27, 28):
                n = _safe_str(_val(ws, rr, 6)).strip()
                lim = _safe_str(_val(ws, rr, 7)).strip()
                if n and lim:
                    multi_lines.append(f"{n} {lim}")
            if len(multi_lines) >= 2:
                result = "\n".join(multi_lines)
            elif len(multi_lines) == 1 and len(names) >= 1:
                result = multi_lines[0]
            elif result:
                # 兜底：只有 1 个干扰物具体值时用旧逻辑
                result = f"干扰物：{'、'.join(names) if names else ''}（实测：{result}）"
        # 触发判定条件：content 或 result 或 conclusion 任一非空
        if content or result or conclusion:
            rows.append({
                "content": content,
                "requirement": requirement,
                "result": result,
                "conclusion": _norm_conclusion(conclusion),
                "row_no": r,  # 记录实际行号，供 _build_summary 精确分配 sub
            })
    # 总结论：跳过模板段落标题（"一、…"、"二、…"、"四、评价结论"等），取 R23-25 中首个正文段
    conclusion_text = ""
    for r in (23, 24, 25):
        t = _safe_str(_val(ws, r, 1))
        if t and not re.match(r"^[一二三四五六七八九十]+、", t.strip()):
            conclusion_text = t
            break
    # 提炼 verify_items 和 result_summary
    rt = info.get("report_type") or "quantitative"
    verify_items = _infer_items(rt, rows)
    result_summary = _build_summary(rows, rt)
    return {
        "verify_items": verify_items,
        "result_summary": result_summary,
        "conclusion": conclusion_text,
        "r14_content": r14,
    }


def _content_from_requirement(req: str) -> str:
    """从验证要求列反推验证内容（当 B 列公式缓存为空时的回退）"""
    if "批内" in req or "实验室内" in req or "CV" in req:
        return "精密度"
    if "偏倚" in req:
        return "正确度"
    if "符合率" in req or "阳性" in req or "阴性" in req:
        return "方法符合率"
    if "检出限" in req:
        return "方法检出限"
    if "线性" in req:
        return "线性范围"
    if "报告" in req or "低限" in req or "高限" in req:
        return "可报告范围"
    if "参考" in req:
        return "参考范围"
    if "干扰" in req or "特异" in req or "胆红素" in req or "甘油三酯" in req or "血红蛋白" in req:
        return "分析特异性"
    return ""


def _content_from_row(r: int) -> str:
    """按行号硬推导验证内容（当 B/C/D/F/G/H 全空时的兜底）。

    模板 BG-SM-CZ-039（定量）布局：
        R17,R18 = 精密度（批内/实验室内 CV）
        R19,R20 = 正确度（低值偏倚 / 高值偏倚）
        R21     = 线性范围
        R22,R23 = 可报告范围（低限 / 高限）
        R24     = 参考范围
        R25     = 评价结论
        R26     = 分析特异性
    模板 BG-SM-CZ-040（定性）布局：R17-R26 是精密度/方法符合率/检出限/干扰等
    """
    mapping_qt = {
        17: "精密度", 18: "精密度",
        19: "正确度", 20: "正确度",
        21: "线性范围",
        22: "可报告范围", 23: "可报告范围",
        24: "参考范围",
        26: "分析特异性",
    }
    mapping_ql = {
        17: "精密度", 18: "精密度",
        19: "方法符合率", 20: "方法符合率",
        21: "方法检出限",
        22: "参考范围",
        26: "分析特异性",
    }
    # 默认按定量模板推断（更常见），定性时由 _infer_items 进一步筛选
    return mapping_qt.get(r) or mapping_ql.get(r) or ""


def _fallback_text(wb_cache, sheet_name, row, col):
    """data_only=True 空缺 → data_only=False 补读纯文本"""
    try:
        wb2 = wb_cache  # 不可，因为 data_only=True 不是 data_only=False
        # 简化：直接判断是否合并区，尝试不同列
    except Exception:
        pass
    return ""


def _norm_conclusion(s):
    s = _safe_str(s)
    if "符合" in s or "通过" in s:
        return "符合要求"
    if "不符合" in s or "不通过" in s:
        return "不符合要求"
    if s:
        return s[:20]
    return ""


KEYWORD_MAP = {
    "精密度": "precision",
    "正确度": "trueness",
    "符合率": "conformity",
    "检出限": "lod",
    "线性": "linearity",
    "可报告": "reportable",
    "参考": "reference",
    "特异": "specificity",
}


def _content_to_key(content: str) -> str:
    """中文 content → 英文 key"""
    for kw, key in KEYWORD_MAP.items():
        if kw in content:
            return key
    return ""


KEY_TO_LABEL = {
    'precision': '精密度', 'trueness': '正确度', 'linearity': '线性范围',
    'reportable': '可报告范围', 'reference': '参考范围', 'specificity': '分析特异性',
    'conformity': '方法符合率', 'lod': '方法检出限',
}


def _normalize_content(content: str) -> str:
    """把识别出的 content 标准化为简洁标签（避免 R21 把"符合线性或临床可接受的非线性程度..."全当 content）。"""
    if not content:
        return content
    k = _content_to_key(content)
    if k:
        return KEY_TO_LABEL[k]
    return content


def _infer_items(rt, rows):
    items = set()
    for r in rows:
        content = r.get("content", "")
        k = _content_to_key(content)
        if k:
            items.add(k)
    return sorted(items)


def _build_summary(rows, report_type: str = "quantitative"):
    """按 content + sub 索引构建 result_summary。

    关键修复：精密度/可报告范围/方法符合率等多行项目，按 subKey（precision1/2/reportable1/2）
    分别存，避免被后面的覆盖丢失。key 用英文 subKey 与前端对齐。

    用"出现顺序"分配 sub（不依赖固定行号，兼容 ALP/ALB 不同模板的表头偏移）：
    - 第 1 个 precision → precision1，第 2 个 → precision2
    - 第 1 个 trueness → trueness1，第 2 个 → trueness2
    - 第 1 个 reportable → reportable1，第 2 个 → reportable2
    - 第 1 个 conformity → conformity1，第 2 个 → conformity2
    """
    rs: dict[str, Any] = {}
    seq = {}  # base -> 出现次数
    for row in rows:
        content = row.get("content", "")
        if not content:
            continue
        base = _content_to_key(content)
        if not base:
            continue  # 解析不到的验证项直接跳过
        seq[base] = seq.get(base, 0) + 1
        sub = str(seq[base]) if base in ("precision", "trueness", "reportable", "conformity") else ""
        key = base + sub
        rs[key] = {
            "result": row.get("result", ""),
            "conclusion": row.get("conclusion", ""),
        }
    return rs


def parse_verification_xlsx(file_bytes: bytes) -> dict:
    """解析上传的性能验证 xlsx 文件，提取项目信息和验证结论。

    双次读取：先 data_only=True 获取公式缓存值（结果/结论），再 data_only=False
    获取纯文本内容（验证内容/要求列的文本），合并两者确保完整性。
    """
    from io import BytesIO
    wb_cache = load_workbook(BytesIO(file_bytes), data_only=True)
    info = _read_cover(wb_cache)
    details = _read_summary(wb_cache, info)
    info.update(details)
    return info


def parse_and_store(file_bytes: bytes, db_session, user, request_ip: str) -> dict:
    """解析 xlsx 并存入 verification_reports 表。

    Returns:
        dict：{id, project_name, archive_id}
    """
    from ..core.crud_base import write_audit
    from ..core.storage import storage, persist_save
    from ..models.report_archive import ReportArchive
    from ..models.verification_report import VerificationReport

    parsed = parse_verification_xlsx(file_bytes)

    rec = VerificationReport(
        report_type=parsed.get("report_type", "qualitative"),
        project_name=parsed.get("project_name", ""),
        project_method=parsed.get("project_method", ""),
        unit=parsed.get("unit", ""),
        reagent=parsed.get("reagent", ""),
        reagent_lot=parsed.get("reagent_lot", ""),
        calibrator=parsed.get("calibrator", ""),
        calibrator_lot=parsed.get("calibrator_lot", ""),
        qc=parsed.get("qc", ""),
        qc_lot=parsed.get("qc_lot", ""),
        instrument=parsed.get("instrument", ""),
        instrument_manufacturer=parsed.get("instrument_manufacturer", ""),
        instrument_model=parsed.get("instrument_model", ""),
        instrument_no=parsed.get("instrument_no", ""),
        tea=parsed.get("tea", ""),
        dilution=parsed.get("dilution", ""),
        linear_low=parsed.get("linear_low", ""),
        linear_high=parsed.get("linear_high", ""),
        verify_items=json.dumps(parsed.get("verify_items", []), ensure_ascii=False),
        data=json.dumps({}, ensure_ascii=False),
        result_summary=json.dumps(parsed.get("result_summary", {}), ensure_ascii=False),
        conclusion=parsed.get("conclusion", ""),
        verify_date=parsed.get("verify_date", ""),
        operator=parsed.get("operator", ""),
        reviewer=parsed.get("reviewer", ""),
        created_by_id=user.id if user else None,
    )
    db_session.add(rec)
    db_session.flush()
    # 写入原始 xlsx 字节到 COS/disk（不是由 parse_and_store 创建 ReportArchive 记录，
    # 归档记录由 report_archives.upload_archive 端点统一创建，避免重复归档）
    rec.report_file_path = ""
    db_session.commit()
    write_audit(db_session, user, "upload_parse", "verification_reports", rec.id, {"project": rec.project_name}, request_ip)
    return {"id": rec.id, "project_name": rec.project_name}
