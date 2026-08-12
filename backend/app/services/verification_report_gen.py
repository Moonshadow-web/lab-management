"""性能验证报告生成器（模板驱动，openpyxl）。

策略：加载科室标准模板（定性 BG-SM-CZ-040 / 定量 BG-SM-CZ-039），
只填写「主封面信息 + 各验证 sheet 数据区 + 结果汇总结论」，
模板自带公式在 Excel 打开时自动联动计算，保证报告格式与模板 100% 一致。
"""
import io
from pathlib import Path

from openpyxl import load_workbook

TEMPLATE_DIR = Path(__file__).resolve().parents[1] / "templates" / "verification"
QUAL_TEMPLATE = TEMPLATE_DIR / "qualitative_template.xlsx"
QUANT_TEMPLATE = TEMPLATE_DIR / "quantitative_template.xlsx"


def _num(v):
    """转数值；空串/None → ''（保留单元格原样处理为空）。"""
    if v is None or v == "":
        return ""
    try:
        return float(v)
    except (TypeError, ValueError):
        return str(v)


def _num3(v):
    n = _num(v)
    return "" if n == "" else round(n, 3)


def _set(ws, row, col, value):
    """安全写值：合并区中间单元格（MergedCell 只读）跳过，左上角可写。"""
    try:
        ws.cell(row=row, column=col).value = value
    except AttributeError:
        pass


def fill_cover(wb, d):
    """主封面：两模板同坐标（E20-E36）。"""
    cover = wb["主封面"]
    _set(cover, 20, 5, d.get("project_name", ""))
    _set(cover, 22, 5, d.get("reagent", ""))
    _set(cover, 24, 5, d.get("instrument", ""))
    _set(cover, 26, 5, d.get("instrument_manufacturer", ""))
    _set(cover, 28, 5, d.get("instrument_model", ""))
    _set(cover, 30, 5, d.get("instrument_no", ""))
    _set(cover, 32, 5, d.get("operator", ""))
    _set(cover, 34, 5, d.get("reviewer", ""))
    _set(cover, 36, 5, d.get("verify_date", ""))


def _fill_precision_data(wb, d):
    """精密度 sheet：C44:E48 水平1 + H44:K48 水平2（5 天 × 3 次）。"""
    p = wb["精密度"]
    data = (d.get("data") or {}).get("precision") or {}
    levels = data.get("levels") or []
    l1 = levels[0] if len(levels) > 0 else None
    l2 = levels[1] if len(levels) > 1 else None
    for i in range(5):
        if l1 and i < len(l1.get("days", [])):
            row = 44 + i
            day = l1["days"][i]
            for j in range(min(3, len(day))):
                p.cell(row=row, column=3 + j, value=_num3(day[j]))
        if l2 and i < len(l2.get("days", [])):
            row = 44 + i
            day = l2["days"][i]
            cols = [8, 9, 11]  # H/I/K（J 列为模板标签列，留空）
            for j in range(min(3, len(day))):
                p.cell(row=row, column=cols[j], value=_num3(day[j]))
    # 精密度结论（R34 定性/定量均为「符合要求」；标题 R32/33 含项目名）
    for ws in wb.worksheets:
        pass
    conclusion = (d.get("result_summary") or {}).get("precision", {}).get("conclusion", "")
    if conclusion:
        for row in (32, 33, 34):
            for col in range(1, 13):
                try:
                    c = wb["精密度"].cell(row=row, column=col)
                    if c.value and "精密度验证" in str(c.value):
                        c.value = f"{d.get('project_name', '')}精密度验证："
                except AttributeError:
                    pass
        _set(wb["精密度"], 34, 1, conclusion)


def _fill_summary(wb, d):
    """结果汇总：参数区 + 验证内容 + 各验证项结果/结论。"""
    s = wb["结果汇总"]
    rtype = d.get("report_type", "qualitative")
    rs = d.get("result_summary") or {}

    # 参数区（B 列左栏 / G 列右栏）
    _set(s, 3, 2, d.get("project_name", ""))
    _set(s, 4, 2, d.get("project_method", ""))
    _set(s, 5, 2, d.get("unit", ""))
    _set(s, 6, 2, d.get("reagent", ""))
    _set(s, 7, 2, d.get("calibrator", ""))
    _set(s, 8, 2, d.get("qc", ""))
    _set(s, 9, 2, d.get("operator", ""))
    _set(s, 10, 2, d.get("verify_date", ""))
    _set(s, 3, 7, d.get("instrument_model", ""))
    _set(s, 4, 7, d.get("instrument_no", ""))
    _set(s, 5, 7, d.get("tea", ""))
    _set(s, 6, 7, d.get("reagent_lot", ""))
    _set(s, 7, 7, d.get("calibrator_lot", ""))
    _set(s, 8, 7, d.get("qc_lot", ""))
    _set(s, 9, 7, d.get("dilution", ""))
    _set(s, 10, 7, d.get("reviewer", ""))
    _set(s, 11, 7, d.get("verify_date", ""))

    # 验证内容（R13 标题 + R14 内容）
    items = d.get("verify_items") or []
    name_map = {
        "precision": "精密度",
        "conformity": "方法符合率",
        "lod": "方法检出限",
        "specificity": "分析特异性",
        "reference": "参考范围",
        "trueness": "正确度",
        "linearity": "线性范围",
        "reportable": "可报告范围",
    }
    content = "、".join([name_map.get(i, i) for i in items if i in name_map])
    if content:
        _set(s, 14, 2, content + "。")

    # 各验证项结果与结论行：按行逐一填充（precision1/precision2 由 _auto_fill_result_summary 填入）
    entries = [
        (17, "precision1"), (18, "precision2"),  # 精密度 批内/实验室内
        (19, "conformity1"), (20, "conformity2"), # 定性 符合率
        (21, "lod"), (21, "trueness"),            # 检出限/正确度 同 R21
        (21, "linearity"),                        # 线性范围 R21
        (22, "reportable1"), (23, "reportable2"), # 可报告 低/高
        (24, "reference"),                        # 参考范围
        (26, "specificity"),                      # 分析特异性
    ]
    for r, sub in entries:
        item = rs.get(sub) or {}
        # 回退到自动匹配主键（如 sub="trueness"，和 rs.trueness 一致）
        text = item.get("result", "")
        concl = item.get("conclusion", "")
        if text:
            _set(s, r, 6, text)   # F 列（验证结果文本）
        if concl:
            _set(s, r, 8, concl)  # H 列（验证结论）
    # 评价结论
    if d.get("conclusion"):
        for r in (24, 25):
            _set(s, r, 1, d["conclusion"])


def _fill_qualitative(wb, d):
    data = d.get("data") or {}
    rs = d.get("result_summary") or {}

    # ---- 方法符合率 ----
    conf = data.get("conformity") or {}
    samples = conf.get("samples") or []
    if samples:
        ws = wb["方法符合率"]
        for i, smp in enumerate(samples[:20]):
            r = 20 + i
            _set(ws, r, 2, smp.get("name", ""))       # B 样品编号
            _set(ws, r, 4, smp.get("ref", ""))        # D 参考判定
            _set(ws, r, 6, _num(smp.get("method", "")))  # F 方法结果值
            _set(ws, r, 8, smp.get("mresult", ""))    # H 方法判定
        # 结论
        concl = (rs.get("conformity") or {}).get("conclusion", "")
        if concl:
            _set(ws, 45, 1, f"{d.get('project_name', '')}方法符合率验证：")
            _set(ws, 46, 1, concl)
    # 原始数据同步
    raw = wb["原始数据"]
    if samples:
        for i, smp in enumerate(samples[:20]):
            r = 36 + i
            _set(raw, r, 2, smp.get("ref", ""))
            _set(raw, r, 4, _num(smp.get("method", "")))

    # ---- 方法检出限 ----
    lod = data.get("lod") or {}
    lod_samples = lod.get("samples") or []
    if lod_samples:
        ws = wb["方法检出限（合适时）"]
        for i, smp in enumerate(lod_samples[:20]):
            r = 20 + i
            _set(ws, r, 2, smp.get("orig", ""))       # B 原浓度
            _set(ws, r, 4, smp.get("diluted", ""))    # D 稀释后浓度
            _set(ws, r, 6, _num(smp.get("value", "")))  # F 结果值
            _set(ws, r, 8, smp.get("mresult", ""))    # H 判定
        concl = (rs.get("lod") or {}).get("conclusion", "")
        if concl:
            _set(ws, 44, 1, f"{d.get('project_name', '')}方法检出限验证：")
            _set(ws, 45, 1, concl)
    if lod_samples:
        for i, smp in enumerate(lod_samples[:20]):
            r = 59 + i
            _set(raw, r, 2, smp.get("orig", ""))
            _set(raw, r, 4, smp.get("diluted", ""))
            _set(raw, r, 6, _num(smp.get("value", "")))

    # ---- 分析特异性（原始数据文本）----
    spec = (data.get("specificity") or {}).get("note", "")
    if spec:
        _set(raw, 82, 1, spec)

    # ---- 原始数据：基本参数 + 精密度 ----
    _set(raw, 3, 2, d.get("project_name", ""))
    _set(raw, 4, 2, d.get("reagent", ""))
    _set(raw, 5, 2, d.get("instrument_model", ""))
    _set(raw, 6, 2, d.get("linear_low", ""))
    _set(raw, 3, 6, d.get("unit", ""))
    _set(raw, 4, 6, d.get("project_method", ""))
    _set(raw, 5, 6, d.get("instrument_no", ""))
    _set(raw, 6, 6, d.get("verify_date", ""))
    _set(raw, 13, 2, d.get("operator", ""))
    _set(raw, 13, 6, d.get("reviewer", ""))
    raw["C15"] = d.get("reviewer", "")
    # 质控靶值
    lv = (data.get("precision") or {}).get("levels") or []
    if len(lv) > 0:
        _set(raw, 16, 2, _num(lv[0].get("target", "")))  # B16 水平1靶值
    if len(lv) > 1:
        _set(raw, 16, 3, _num(lv[1].get("target", "")))  # C16 水平2靶值
    # 精密度原始数据
    for i in range(5):
        if len(lv) > 0 and i < len(lv[0].get("days", [])):
            day = lv[0]["days"][i]
            for j in range(min(3, len(day))):
                _set(raw, 21 + i, 2 + j, _num3(day[j]))
        if len(lv) > 1 and i < len(lv[1].get("days", [])):
            day = lv[1]["days"][i]
            for j in range(min(3, len(day))):
                _set(raw, 28 + i, 2 + j, _num3(day[j]))


def _fill_quantitative(wb, d):
    data = d.get("data") or {}
    rs = d.get("result_summary") or {}

    # ---- 正确度-偏倚评估：R40-R44（5 天 × C 第1次 + E 第2次）----
    tru = data.get("trueness") or {}
    levels = tru.get("levels") or []
    if levels:
        ws = wb["正确度-偏倚评估"]
        l1 = levels[0]
        days = l1.get("days", [])
        for i in range(min(5, len(days))):
            r = 40 + i
            day = days[i]
            if len(day) > 0:
                _set(ws, r, 3, _num3(day[0]))  # C 第1次
            if len(day) > 1:
                _set(ws, r, 5, _num3(day[1]))  # E 第2次
        if l1.get("target"):
            _set(ws, 15, 5, _num(l1["target"]))  # E15 标准物质赋值
        concl = (rs.get("trueness") or {}).get("conclusion", "")
        if concl:
            _set(ws, 34, 1, f"{d.get('project_name', '')}正确度验证：")
            _set(ws, 35, 1, concl)
    # 原始数据联动（R38-R44 引用正确度 sheet，自动带出；不需填）

    # ---- 线性范围：稀释比例 C37:F38 + 数据 R42-R47（B/C/D 3 次）----
    lin = data.get("linearity") or {}
    points = lin.get("points") or []
    if points:
        ws = wb["线性范围"]
        lows = lin.get("low_ratios") or []
        highs = lin.get("high_ratios") or []
        for i in range(min(6, len(lows))):
            _set(ws, 37, 3 + i, _num(lows[i]))
        for i in range(min(6, len(highs))):
            _set(ws, 38, 3 + i, _num(highs[i]))
        for i in range(min(6, len(points))):
            r = 42 + i
            p = points[i]
            _set(ws, r, 2, _num3(p.get("v1", "")))
            _set(ws, r, 3, _num3(p.get("v2", "")))
            _set(ws, r, 4, _num3(p.get("v3", "")))
        concl = (rs.get("linearity") or {}).get("conclusion", "")
        if concl:
            _set(ws, 53, 1, concl)

    # ---- 原始数据：基本参数 + 精密度（原始数据自动引用精密度 sheet，不需手填）----
    raw = wb["原始数据"]
    _set(raw, 4, 2, d.get("reagent", ""))        # R4C2 试剂品牌
    _set(raw, 14, 2, d.get("operator", ""))

    # ---- 可报告范围/可报告范围验证 sheet 结论 ----
    rp = (rs.get("reportable") or {})
    if rp.get("conclusion") or rp.get("result"):
        for sn in ("可报告范围", "可报告范围验证"):
            ws = wb[sn]
            text = f"{d.get('project_name', '')}可报告范围：{rp.get('result', '')}"
            _set(ws, 26, 1, text)
            _set(ws, 27, 1, rp.get("conclusion", ""))

    # ---- 参考区间 sheet 结论 ----
    ref = (rs.get("reference") or {})
    if ref.get("conclusion") or ref.get("result"):
        ws = wb["参考区间"]
        _set(ws, 59, 1, f"{d.get('project_name', '')}参考区间验证通过，参考区间为：{ref.get('result', '')}")

    # ---- 分析特异性 sheet 结论 ----
    spec = (rs.get("specificity") or {})
    if spec.get("conclusion") or spec.get("result"):
        ws = wb["分析特异性"]
        _set(ws, 26, 1, f"{d.get('project_name', '')}分析特异性验证：{spec.get('result', '')}")
        _set(ws, 27, 1, spec.get("conclusion", ""))
    _set(raw, 14, 6, d.get("reviewer", ""))


def build_verification_report(d: dict) -> bytes:
    """根据记录数据生成 xlsx 报告（模板驱动），返回字节。"""
    rtype = d.get("report_type", "qualitative")
    template = QUAL_TEMPLATE if rtype == "qualitative" else QUANT_TEMPLATE
    if not template.exists():
        raise FileNotFoundError(f"模板缺失：{template}")
    wb = load_workbook(template)  # data_only=False，公式保留
    fill_cover(wb, d)
    _fill_precision_data(wb, d)
    _fill_summary(wb, d)
    if rtype == "qualitative":
        _fill_qualitative(wb, d)
    else:
        _fill_quantitative(wb, d)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()