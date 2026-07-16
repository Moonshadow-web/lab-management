"""EQA v2 验证：上报逻辑、报告导入/下载/删除、导出表头、通知文案。"""
import urllib.request
import urllib.parse
import json
import datetime
import os
import struct

BASE = "http://127.0.0.1:8123"
TODAY = datetime.date.today()
DUE = (TODAY + datetime.timedelta(days=10)).strftime("%Y-%m-%d")  # 临期 → warning
OUT = os.path.join(os.path.dirname(__file__), "sample_report.pdf")

passed = 0
failed = 0


def check(name, ok, extra=""):
    global passed, failed
    if ok:
        passed += 1
        print(f"[PASS] {name}")
    else:
        failed += 1
        print(f"[FAIL] {name}  {extra}")


def req(method, path, data=None, form=False, raw=False, headers=None):
    url = BASE + path
    h = headers or {}
    body = None
    if data is not None:
        if isinstance(data, bytes):
            body = data  # 原始字节（multipart），不 JSON 编码
        elif form:
            body = urllib.parse.urlencode(data).encode()
            h["Content-Type"] = "application/x-www-form-urlencoded"
        else:
            body = json.dumps(data).encode()
            h["Content-Type"] = "application/json"
    r = urllib.request.Request(url, data=body, method=method, headers=h)
    try:
        with urllib.request.urlopen(r, timeout=15) as resp:
            if raw:
                return resp.status, resp.read(), resp.headers
            ct = resp.headers.get("Content-Type", "")
            if "application/json" in ct:
                return resp.status, json.loads(resp.read().decode()), resp.headers
            return resp.status, resp.read(), resp.headers
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode(), e.headers


# 登录
s, b, _ = req("POST", "/api/v1/auth/login",
              {"username": "admin", "password": "admin123"}, form=True)
check("登录", s == 200 and "access_token" in b, f"status={s}")
TOKEN = b["access_token"]
AH = {"Authorization": f"Bearer {TOKEN}"}

# 创建一个临期未上报计划
plan = {"year": 2026, "org": "验证机构", "program": "生化", "item": "葡萄糖",
        "round_no": "R9", "sample_date": TODAY.strftime("%Y-%m-%d"), "due_date": DUE,
        "returned": False, "result": "", "qualified": False, "score": "", "note": ""}
s, b, _ = req("POST", "/api/v1/eqa-plans", plan, headers=AH)
check("创建计划(201)", s == 201 and isinstance(b, dict) and "id" in b, f"status={s}")
pid = b.get("id")

# alerts 应包含它（warning）
s, b, _ = req("GET", "/api/v1/eqa-plans/alerts", headers=AH)
alert_ids = [a["plan_id"] for a in b]
check("检测提醒含临期计划", pid in alert_ids, f"alerts={alert_ids}")
# 注：/alerts 仅返回结构化数据，前端渲染「上报截止」文案（已在 QCList.vue 改为 上报截止）

# 标记已上报
s, b, _ = req("PUT", f"/api/v1/eqa-plans/{pid}", {**plan, "returned": True}, headers=AH)
check("标记已上报(200)", s == 200 and b.get("returned") is True, f"status={s}")
s, b, _ = req("GET", "/api/v1/eqa-plans/alerts", headers=AH)
check("已上报后退出提醒", pid not in [a["plan_id"] for a in b], f"alerts={b}")

# 首页通知文案：室间质评上报提醒
s, b, _ = req("GET", "/api/v1/notifications", headers=AH)
items = b.get("items", []) if isinstance(b, dict) else b
eqa_notes = [n for n in items if n.get("ref_type") == "eqa_return"]
check("通知含 室间质评上报提醒", any("室间质评上报提醒" in (n.get("title") or "") for n in eqa_notes),
      f"titles={[n.get('title') for n in eqa_notes][:3]}")

# 造一个最小合法 PDF
pdf = b"%PDF-1.4\n1 0 obj<</Type/Catalog>>endobj\ntrailer<</Root 1 0 R>>\n%%EOF"
with open(OUT, "wb") as f:
    f.write(pdf)

# 导入报告（multipart）
boundary = "----eqaboundary"
raw_body = (
    f"--{boundary}\r\n"
    f'Content-Disposition: form-data; name="file"; filename="sample_report.pdf"\r\n'
    f"Content-Type: application/pdf\r\n\r\n"
).encode() + pdf + f"\r\n--{boundary}--\r\n".encode()
s, b, _ = req("POST", f"/api/v1/eqa-plans/report/{pid}",
              data=raw_body, headers={**AH, "Content-Type": f"multipart/form-data; boundary={boundary}"})
check("导入报告(200)", s == 200 and b.get("report_file", "").endswith(".pdf"), f"status={s} body={b}")
rep_path = b.get("report_file")

# 下载报告
s, b, _ = req("GET", f"/api/v1/eqa-plans/report/{pid}", raw=True, headers=AH)
check("下载报告(PDF头)", s == 200 and b[:4] == b"%PDF", f"status={s} head={b[:8]}")

# 非PDF 应拒绝
bad = b"not a pdf"
boundary2 = "----eqaboundary2"
raw_bad = (
    f"--{boundary2}\r\n"
    f'Content-Disposition: form-data; name="file"; filename="x.txt"\r\n'
    f"Content-Type: text/plain\r\n\r\n"
).encode() + bad + f"\r\n--{boundary2}--\r\n".encode()
s, b, _ = req("POST", f"/api/v1/eqa-plans/report/{pid}",
              data=raw_bad, headers={**AH, "Content-Type": f"multipart/form-data; boundary={boundary2}"})
check("拒绝非PDF(400)", s == 400, f"status={s}")

# 导出 Excel 表头
import io
s, b, h = req("GET", "/api/v1/eqa-plans/export?year=2026", raw=True, headers=AH)
is_xlsx = b[:2] == b"PK"
cd = h.get("Content-Disposition", "")
check("导出为xlsx", is_xlsx and "UTF-8" in cd, f"status={s} head={b[:4]} cd={cd}")
if is_xlsx:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(b))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        check("导出表头含 上报截止日期/是否上报/报告",
              ("上报截止日期" in headers) and ("是否上报" in headers) and ("报告" in headers),
              f"headers={headers}")
    except Exception as e:
        check("导出表头解析", False, str(e))

# 删除报告
s, b, _ = req("DELETE", f"/api/v1/eqa-plans/report/{pid}", headers=AH)
check("删除报告(200)", s == 200, f"status={s}")
s, b, _ = req("GET", f"/api/v1/eqa-plans/report/{pid}", raw=True, headers=AH)
check("删除后下载404", s == 404, f"status={s}")

# 清理测试计划
s, b, _ = req("DELETE", f"/api/v1/eqa-plans/{pid}", headers=AH)
check("清理测试计划", s == 200, f"status={s}")

if os.path.exists(OUT):
    os.remove(OUT)

print(f"\n=== 结果：通过 {passed}，失败 {failed} ===")
