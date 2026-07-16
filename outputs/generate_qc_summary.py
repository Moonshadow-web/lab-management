import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

QUALITY_GOALS = {
    "甲胎蛋白": "7.5%",
    "人绒毛膜促性腺激素": "8.3%",
    "游离雌三醇": "",
}

# 手工从 9 页 PDF 提取的汇总数据（每页 = 一个项目/季度）
PAGES = [
    # (项目, 单位, 季度, 月份范围, 仪器, 水平数据)
    ("甲胎蛋白", "ug/L", "Q1", "2025-01-01 至 2025-03-31", "DXI800", [
        (1, "1002101", 11.1, 0.55, "4.95%", 10.95, 0.4932, "4.50%", 12, 1),
        (2, "1002102", 33.8, 1.67, "4.34%", 33.1877, 1.3343, "4.20%", 12, 1),
        (3, "1002103", 80, 3.8, "4.75%", 77.4531, 2.4862, "3.18%", 12, 1),
    ]),
    ("游离雌三醇", "ug/L", "Q1", "2025-01-01 至 2025-03-31", "DXI800", [
        (1, "1002101", 0.535, 0.04, "7.40%", 0.5322, 0.0423, "7.95%", 12, 1),
        (2, "1002102", 1.77, 0.14, "7.91%", 1.7875, 0.1301, "7.28%", 12, 1),
        (3, "1002103", 3.3, 0.26, "7.88%", 3.3402, 0.2512, "7.70%", 12, 1),
    ]),
    ("人绒毛膜促性腺激素", "IU/L", "Q1", "2025-01-01 至 2025-03-31", "DXI800", [
        (1, "1002101", 11270, 550, "4.88%", 11256.98, 414.7006, "3.68%", 12, 0),
        (2, "1002102", 38034, 1856, "4.88%", 37890.42, 1270.1701, "3.35%", 12, 0),
        (3, "1002103", 75582, 3400, "4.50%", 78734.14, 3663.7871, "4.65%", 12, 0),
    ]),
    ("甲胎蛋白", "ug/L", "Q2", "2025-04-01 至 2025-06-30", "DXI800", [
        (1, "1002101", 11.1, 0.55, "4.95%", 11.1277, 0.4434, "4.01%", 12, 3),
        (2, "1002102", 33.3, 1.66, "4.98%", 33.2025, 1.4645, "4.41%", 12, 3),
        (3, "1002103", 77, 3.8, "4.94%", 78.235, 3.8714, "4.94%", 12, 4),
    ]),
    ("游离雌三醇", "ug/L", "Q2", "2025-04-01 至 2025-06-30", "DXI800", [
        (1, "1002101", 0.535, 0.04, "7.46%", 0.5548, 0.038, "6.85%", 12, 1),
        (2, "1002102", 1.77, 0.14, "7.91%", 1.8319, 0.1102, "6.02%", 12, 1),
        (3, "1002103", 3.3, 0.26, "7.88%", 3.5288, 0.2213, "6.27%", 12, 1),
    ]),
    ("人绒毛膜促性腺激素", "IU/L", "Q2", "2025-04-01 至 2025-06-30", "DXI800", [
        (1, "1002101", 11270, 550, "4.88%", 11256.98, 414.7006, "3.68%", 12, 0),
        (2, "1002102", 38034, 1856, "4.88%", 37432.83, 1270.1701, "3.39%", 12, 0),
        (3, "1002103", 75582, 3400, "4.50%", 76596, 2594.9487, "3.39%", 12, 0),
    ]),
    ("甲胎蛋白", "ug/L", "Q3", "2025-07-01 至 2025-09-30", "DXI800", [
        (1, "1002101", 11.1, 0.55, "4.95%", 11.0871, 0.5423, "4.89%", 12, 1),
        (2, "1002102", 33.3, 1.66, "4.98%", 33.3223, 1.5794, "4.74%", 12, 2),
        (3, "1002103", 77, 3.8, "4.94%", 78.8262, 3.8919, "4.94%", 12, 3),
    ]),
    ("游离雌三醇", "ug/L", "Q3", "2025-07-01 至 2025-09-30", "DXI800", [
        (1, "1002101", 0.545, 0.04, "7.34%", 0.58, 0.0493, "8.50%", 12, 1),
        (2, "1002102", 1.77, 0.14, "7.91%", 1.8801, 0.1193, "6.35%", 12, 1),
        (3, "1002103", 3.4, 0.26, "7.65%", 3.4686, 0.1939, "5.59%", 12, 1),
    ]),
    ("人绒毛膜促性腺激素", "IU/L", "Q3", "2025-07-01 至 2025-09-30", "DXI800", [
        (1, "1002101", 11270, 550, "4.88%", 11650.64, 427.2071, "3.67%", 12, 0),
        (2, "1002102", 38034, 1856, "4.88%", 38926.38, 1381.2641, "3.55%", 12, 0),
        (3, "1002103", 75582, 3400, "4.50%", 77467, 2627.4424, "3.39%", 12, 0),
    ]),
]


def parse_pct(p):
    try:
        return float(p.replace("%", "")) / 100
    except Exception:
        return None


def build_excel(out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "室内质控月小结"

    # 标题
    ws.merge_cells("A1:T1")
    ws["A1"] = "生化免疫组室内质控月小结"
    ws["A1"].font = Font(name="宋体", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 信息行
    ws.merge_cells("A2:T2")
    ws["A2"] = "仪器：特种蛋白分析仪 DXI800      质控批号：1002101/1002102/1002103      年：2025"
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells("A3:T3")
    ws["A3"] = "本月仪器情况总结：仪器运行良好；共统计 2025 年 Q1-Q3 甲胎蛋白、游离雌三醇、人绒毛膜促性腺激素三个项目的室内质控数据。"
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 36

    # 表头
    headers = [
        "项目", "质控批号", "单位", "水平", "季度/月份",
        # 靶值
        "靶值", "靶值SD", "靶值CV%",
        # 当月原始测定结果
        "均值", "SD", "CV%", "n", "失控数", "在控率",
        # 剔除失控后的累积测定结果
        "均值", "SD", "CV%", "n", "在控率",
        # 质量目标
        "质量目标（允许不精密度）",
    ]
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill = PatternFill("solid", fgColor="E7E6E6")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, size=10)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 36

    row = 5
    for project, unit, quarter, date_range, instrument, levels in PAGES:
        qgoal = QUALITY_GOALS.get(project, "")
        for level, lot, target, target_sd, target_cv, obs_mean, obs_sd, obs_cv, n, out_count in levels:
            in_control = n - out_count
            in_control_rate = in_control / n if n else None
            # 如无失控，将剔除失控后的累积结果暂用原始测定结果；有待逐日值精确计算
            if out_count == 0:
                cum_mean, cum_sd, cum_cv, cum_n, cum_rate = obs_mean, obs_sd, obs_cv, n, in_control_rate
            else:
                cum_mean = cum_sd = cum_cv = cum_n = cum_rate = None

            values = [
                project, lot, unit, level, f"{quarter}（{date_range}）",
                target, target_sd, target_cv,
                obs_mean, obs_sd, obs_cv, n, out_count, in_control_rate,
                cum_mean, cum_sd, cum_cv, cum_n, cum_rate,
                qgoal,
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col != 1 else "left", vertical="center", wrap_text=True)
            row += 1

    # 合计/审批行
    ws.merge_cells(f"A{row}:T{row}")
    ws.cell(row=row, column=1, value="失控及处理情况总结：详见各季度质控说明，主要失控原因包括定标时间过长、质控液被污染等，均已按规范处理并重新定标。")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"A{row}:T{row}")
    ws.cell(row=row, column=1, value="质控总负责人审批意见：")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    row += 1

    ws.merge_cells(f"A{row}:T{row}")
    ws.cell(row=row, column=1, value="审批人签字：__________    日期：__________")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

    # 列宽
    widths = [12, 12, 8, 6, 28, 10, 10, 10, 10, 10, 8, 6, 8, 10, 10, 10, 8, 6, 10, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 百分比格式
    for r in range(5, row - 3):
        for c in [13, 19]:  # 在控率
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"

    wb.save(out_path)
    print(f"Saved {out_path}")


if __name__ == "__main__":
    build_excel(r"D:/workbuddyprojects/网页版-生免速查工具/outputs/室内质控月小结_2025_Q1-Q3.xlsx")
