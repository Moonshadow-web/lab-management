import urllib.request, urllib.parse, json, os

BASE = "http://127.0.0.1:8123"
CSV_PATH = r"d:\workbuddyprojects\网页版-生免速查工具\outputs\test_qc_jan2025.csv"
EXPORT_PATH = r"d:\workbuddyprojects\网页版-生免速查工具\outputs\test_export.xlsx"


def login():
    data = urllib.parse.urlencode({"username": "admin", "password": "admin123"}).encode()
    req = urllib.request.Request(BASE + "/api/v1/auth/login", data=data,
                                 headers={"Content-Type": "application/x-www-form-urlencoded"})
    return json.loads(urllib.request.urlopen(req).read())["access_token"]


def req_json(url, token, method="GET", data=None, headers=None):
    h = {"Authorization": "Bearer " + token}
    if headers:
        h.update(headers)
    r = urllib.request.Request(url, data=data, method=method, headers=h)
    return json.loads(urllib.request.urlopen(r).read())


def upload_csv(token):
    boundary = "----qcboundary123456"
    with open(CSV_PATH, "rb") as f:
        filedata = f.read()
    parts = [
        ("--" + boundary).encode(),
        b'Content-Disposition: form-data; name="file"; filename="test_qc_jan2025.csv"',
        b"Content-Type: text/csv",
        b"",
        filedata,
        ("--" + boundary + "--").encode(),
        b"",
    ]
    content = b"\r\n".join(parts)
    h = {"Content-Type": "multipart/form-data; boundary=" + boundary}
    return req_json(BASE + "/api/v1/qc-summaries/upload", token, method="POST", data=content, headers=h)


def main():
    token = login()
    print("LOGIN OK")
    up = upload_csv(token)
    print("UPLOAD:", json.dumps(up, ensure_ascii=False)[:500])

    lst = req_json(BASE + "/api/v1/qc-summaries?year=2025&month=1&page_size=100", token)
    items = lst["items"]
    print(f"\nLIST total={lst['total']}, returned={len(items)}")
    for it in items:
        print(f"  [{it['id']}] {it['test_item']} L{it['level']} 批{it['lot_no']} "
              f"n={it['n']} 失控={it['out_of_control_count']} 在控率={it['in_control_rate']} "
              f"质量目标={it['quality_goal']}")

    # daily for first item
    if items:
        sid = items[0]["id"]
        daily = req_json(BASE + f"/api/v1/qc-summaries/{sid}/daily", token)
        ooc = [d for d in daily if d["is_out_of_control"]]
        print(f"\nDAILY for id={sid}: {len(daily)} 条, 失控点 {len(ooc)}:")
        for d in ooc:
            print(f"    {d['qc_date']} val={d['value']} rule={d['rule_violated']}")

    # export (returns binary Excel; save + verify)
    h = {"Authorization": "Bearer " + token}
    r = urllib.request.Request(BASE + "/api/v1/qc-summaries/export?year=2025&month=1", headers=h)
    body = urllib.request.urlopen(r).read()
    with open(EXPORT_PATH, "wb") as f:
        f.write(body)
    print(f"\nEXPORT saved: {EXPORT_PATH} ({len(body)} bytes)")
    # verify content via openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(EXPORT_PATH)
    ws = wb.active
    print(f"  sheets={wb.sheetnames}, max_row={ws.max_row}, max_col={ws.max_column}")
    # print first data row (14 cols) after title/header
    for rr in range(1, ws.max_row + 1):
        row = [ws.cell(row=rr, column=c).value for c in range(1, 15)]
        if any(v not in (None, "") for v in row):
            print(f"  row{rr}: {row}")


if __name__ == "__main__":
    main()
