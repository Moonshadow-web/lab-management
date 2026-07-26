"""从线上 API 拉取仪器耗材关联数据，按总型号分组，导出 Excel。"""
import json
import re
import sys
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


def extract_base_model(name: str, model: str = "") -> str:
    """与后端 _extract_instrument_base_model 完全一致（含家族映射表）。"""
    # 家族映射表（用户 2026-07-26 确认）
    FAMILY_MAP = {
        "AU5800": "AU58",
        "TOP700A": "TOP700",
        "TOP700B": "TOP700",
        "TOP700C": "TOP700",
        "免疫分析仪": "罗氏Cobas6000",
        "DXA5000": None,  # 无耗材，不显示
    }
    if not name:
        return (model or "").strip()
    raw = name.strip()
    if raw in FAMILY_MAP:
        return FAMILY_MAP[raw]
    m = re.match(r"^([A-Za-z0-9]+(?:-[A-Za-z0-9]+)*)", raw)
    if m:
        base = m.group(1)
        base = re.sub(r"-\d+$", "", base)
        if not base.isdigit() and len(base) >= 2:
            return FAMILY_MAP.get(base, base)
    if model:
        m2 = re.search(r"([A-Za-z]{2,}\d*[A-Za-z]*)", model)
        if m2 and len(m2.group(1)) >= 2:
            return FAMILY_MAP.get(m2.group(1), m2.group(1))
    return raw


def main():
    # 加载已拉取的原始数据（两页）
    all_items = []
    for fname in [
        "d:/workbuddyprojects/网页版-生免速查工具/inst_p1.json",
        "d:/workbuddyprojects/网页版-生免速查工具/inst_p2.json",
    ]:
        try:
            with open(fname, encoding="utf-8") as f:
                d = json.load(f)
                all_items.extend(d.get("items", []))
        except FileNotFoundError:
            pass

    print(f"原始数据: {len(all_items)} 条")

    # 分组：(base_model, reagent_name) → [rows...]
    grouped = defaultdict(list)
    for it in all_items:
        name = it.get("instrument_name", "")
        model = it.get("instrument_model", "")
        bm = extract_base_model(name, model)
        if not bm:  # 家族映射为 None（如 DXA5000 无耗材）跳过
            continue
        rname = it.get("reagent_name", "")
        grouped[(bm, rname)].append(it)

    # 展平并排序
    rows = []
    for (bm, rname), grp in sorted(grouped.items()):
        inst_names = sorted({it.get("instrument_name", "") for it in grp})
        types_ = sorted({it.get("reagent_type", "") for it in grp})
        libs = sorted({it.get("reagent_library", "") for it in grp})
        rows.append({
            "base_model": bm,
            "instruments": "、".join(inst_names),
            "count": len(grp),
            "reagent": rname,
            "type": ",".join(types_),
            "library": ",".join(libs),
        })
    rows.sort(key=lambda r: (r["base_model"], r["reagent"]))

    # 统计
    models = sorted(set(r["base_model"] for r in rows))
    print(f"\n总型号数: {len(models)}")
    for m in models:
        cnt = sum(1 for r in rows if r["base_model"] == m)
        insts = set()
        for r in rows:
            if r["base_model"] == m:
                insts.update(r["instruments"].split("、"))
        print(f"  {m}: {cnt} 种耗材 / {len(insts)} 台实例")

    # 写 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "仪器耗材对照表"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["序号", "仪器总型号", "包含实例（台）", "台数", "耗材名称", "耗材类型", "责任库"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for i, r in enumerate(rows, 1):
        ws.cell(row=i + 1, column=1, value=i).alignment = center
        ws.cell(row=i + 1, column=2, value=r["base_model"]).alignment = left_align
        ws.cell(row=i + 1, column=3, value=r["instruments"]).alignment = left_align
        ws.cell(row=i + 1, column=4, value=r["count"]).alignment = center
        ws.cell(row=i + 1, column=5, value=r["reagent"]).alignment = left_align
        ws.cell(row=i + 1, column=6, value=r["type"]).alignment = center
        ws.cell(row=i + 1, column=7, value=r["library"]).alignment = center
        for col in range(1, 8):
            ws.cell(row=i + 1, column=col).border = thin_border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 7
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 12

    out = "d:/workbuddyprojects/网页版-生免速查工具/仪器耗材对照表.xlsx"
    wb.save(out)
    print(f"\n✅ 已导出: {out}")


if __name__ == "__main__":
    main()
