# -*- coding: utf-8 -*-
import json, glob, os, datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

ROOT = os.path.dirname(os.path.abspath(__file__))
TODAY = "2026-07-10"

# 1) load pending 81 projects
pending = json.load(open(os.path.join(ROOT, "tmp_pending81.json"), encoding="utf-8"))
# ensure order by id
pending = sorted(pending, key=lambda r: int(r["id"]))

# 2) load sop_out extracted data  {pid: {...}}
ext = {}
for f in glob.glob(os.path.join(ROOT, "sop_out", "*.json")):
    d = json.load(open(f, encoding="utf-8"))
    ext[str(d["pid"])] = d

# 3) load old prefilled Excel (GL pre-fill) -> map by position (row2 = pid125 ...)
old = openpyxl.load_workbook(os.path.join(ROOT, "检验项目_待补清单_预填.xlsx"))
ows = old["待补清单_预填"]
# header
old_hdr = [c.value for c in ows[1]]
# build pid -> old row values
old_by_pid = {}
for ri, row in enumerate(ows.iter_rows(min_row=2, values_only=True), start=2):
    pid = 124 + ri  # row2->125
    old_by_pid[str(pid)] = row

# mapping old column name -> index (0-based)
def ocol(name):
    return old_hdr.index(name)

# styles
green = PatternFill("solid", fgColor="C6EFCE")
yellow = PatternFill("solid", fgColor="FFEB9C")
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(color="FFFFFF", bold=True)
wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# output columns
COLS = ["项目编号","项目名称","别名","类别","标本类型","方法学","单位","参考范围",
        "收费","使用仪器","仪器组","线性范围","稀释倍数","可报告范围","稀释液",
        "校准品","溯源性","最近更新","溶血干扰","胆红素干扰","脂血干扰","数据来源","备注"]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "待补清单_SOP预填"
ws.append(COLS)
for ci, _ in enumerate(COLS, start=1):
    c = ws.cell(1, ci)
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = wrap; c.border = border

NO_REF_TOKENS = {"未列出", "未找到（需人工补充）", "", None}

filled = 0
need_input = 0
for rec in pending:
    pid = str(rec["id"])
    name = rec.get("name") or ""
    aliases = rec.get("aliases") or ""
    cat = rec.get("category") or ""
    spec = rec.get("specimen") or ""
    e = ext.get(pid, {})
    old_row = old_by_pid.get(pid)
    def gv(colname):
        if old_row is None: return None
        try: return old_row[ocol(colname)]
        except Exception: return None

    ref = e.get("reference_range", "") or ""
    # decide ref fill + note
    note = ""
    data_source = "待补"
    if pid in ext:
        ds = e.get("data_source", "")
        data_source = "仪器/试剂说明书" if ds == "说明书" else "项目SOP"
    # normalize ref blank
    ref_fill = ref if ref not in NO_REF_TOKENS else ""
    if ref in ("未列出",):
        note = "定性项目，无数值参考范围（需人工确认）"
    elif ref in ("未找到（需人工补充）",):
        note = "KB无相关手册，需人工补充"
    if pid == "199":
        note = "代理判读(安图HSV-1 IgM SOP：<6阴性/6-10可疑/≥10阳性)，需人工核对"

    # method/unit/instrument: prefer SOP, else GL
    method = e.get("method") or gv("方法学*")
    unit = e.get("unit") or gv("单位*")
    instrument = e.get("instrument") or gv("使用仪器")
    # reportable: prefer SOP reportable_range, else GL 可报告范围*
    rep = e.get("reportable_range") or ""
    if not rep or rep in NO_REF_TOKENS:
        rep = gv("可报告范围*")
    # traceability: prefer SOP if meaningful
    trac = e.get("traceability") or ""
    if not trac or trac in NO_REF_TOKENS:
        trac = gv("溯源性*")

    row = [
        int(pid), name, aliases, cat, spec,
        method, unit, ref_fill,
        gv("收费"), instrument, gv("仪器组"),
        gv("线性范围*"), gv("稀释倍数"), rep, gv("稀释液*"),
        gv("校准品*"), trac, TODAY,
        gv("溶血干扰*"), gv("胆红素干扰*"), gv("脂血干扰*"),
        data_source, note,
    ]
    ws.append(row)
    r = ws.max_row
    for ci in range(1, len(COLS)+1):
        cell = ws.cell(r, ci); cell.alignment = wrap; cell.border = border
    # color 参考范围 cell (col 8)
    rc = ws.cell(r, 8)
    if ref_fill:
        rc.fill = green; filled += 1
    else:
        rc.fill = yellow; need_input += 1

# column widths
widths = [10, 24, 14, 8, 10, 18, 8, 40, 8, 22, 10, 14, 10, 16, 14, 16, 22, 12, 12, 12, 12, 14, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# 填写说明 sheet
ws2 = wb.create_sheet("填写说明")
notes = [
    "本表由系统自动生成，共 81 条（即数据库中参考范围为空的项目）。",
    "本次在上一版 GL 文档预填基础上，新增「参考范围」列，数据来源优先级：",
    "  1) 具体项目 SOP（生免组知识库「项目SOP」文件夹）第11节 生物参考区间 / 第6.1节 参考范围；",
    "  2) 无 SOP 者查仪器/试剂说明书；",
    "  3) 仍缺者留空标黄，请人工补充。",
    "",
    "【颜色含义】",
    "  绿色「参考范围」：已从 SOP/说明书提取，请核对后保留或修正。",
    "  黄色「参考范围」：系统未能提取，需人工补充（见「备注」列说明）。",
    "",
    "【待人工确认项】",
    "  - pid167 免疫固定电泳：定性电泳项目，无数值参考范围。",
    "  - pid199 II型单纯疱疹病毒IgM抗体：KB无临床参考范围，暂以同方法学(安图Autolumo A2000)HSV-1 IgM SOP判读为代理值，需核对。",
    "  - pid202 P2PSA：KB未收录参考范围，需人工补充。",
    "  - pid203 ST2：KB未收录任何手册，需人工补充。",
    "  - pid150/151/154 总蛋白(尿/脑脊液)：对应 SOP 正文疑似错配(实为NAG项目)，参考范围已取正文值，请重点复核。",
    "  - pid156/204 腺苷脱氨酶：SOP 生物参考区间正文为空，参考范围取自结果报告方式表(0-40 U/L)，请复核。",
    "",
    "【其他列】线性范围/可报告范围/稀释/校准品/溯源性/干扰 等来自 GL 文档预填，已尽量用 SOP 同名项覆盖。",
]
for i, t in enumerate(notes, start=1):
    ws2.cell(i, 1, t)
ws2.column_dimensions["A"].width = 100

out = os.path.join(ROOT, "检验项目_待补清单_SOP预填.xlsx")
wb.save(out)
print("saved:", out)
print("total rows:", len(pending))
print("参考范围 已填(绿):", filled, " 待补(黄):", need_input)
