"""生成「试剂批号待回补清单」：当前库存里没有批号的**试剂**（type=试剂）。

按年用量降序，供下次到货时优先补录批号/效期。
用法：python scripts/report_missing_batch.py [--all-types]
"""
import json
import sys
import urllib.parse
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

H = "http://lab-management-282724-9-1408547492.sh.run.tcloudbase.com"
OUT = "outputs/试剂批号待回补清单.xlsx"
HEAD_FILL = PatternFill("solid", fgColor="1a365d")
TOP_FILL = PatternFill("solid", fgColor="FFF2CC")


def login():
    data = urllib.parse.urlencode(
        {"username": "jinzizheng", "password": "Jzz6827556"}).encode()
    req = urllib.request.Request(H + "/api/v1/auth/login", data=data)
    return json.load(urllib.request.urlopen(req, timeout=60))["access_token"]


def get(tok, path):
    req = urllib.request.Request(H + path,
                                 headers={"Authorization": "Bearer " + tok})
    return json.load(urllib.request.urlopen(req, timeout=300))


def paged(tok, path, size=200):
    out, page = [], 1
    while True:
        d = get(tok, f"{path}&page={page}&page_size={size}")
        out += d["items"]
        if len(out) >= d["total"] or not d["items"]:
            break
        page += 1
    return out


def main():
    all_types = "--all-types" in sys.argv
    tok = login()
    items = {i["id"]: i for i in paged(tok, "/api/v1/reagent/items?page=1")}
    stock = paged(tok, "/api/v1/reagent/stock?page=1")

    no_batch = {r["item_id"]: r for r in stock if not (r["batch_no"] or "").strip()}

    rows = []
    for iid, r in no_batch.items():
        it = items.get(iid)
        if not it:
            continue
        if not all_types and it.get("type") != "试剂":
            continue
        rows.append({
            "id": iid, "name": it.get("name", ""), "spec": it.get("spec", ""),
            "brand": it.get("brand", ""), "type": it.get("type", ""),
            "library": it.get("library", ""),
            "material_code": it.get("material_code", ""),
            "unit": it.get("unit", ""),
            "annual_usage": int(it.get("annual_usage") or 0),
            "min_stock": int(it.get("min_stock") or 0),
            "qty": int(r["quantity"] or 0),
        })
    rows.sort(key=lambda x: (-x["annual_usage"], x["name"]))

    wb = Workbook()
    ws = wb.active
    ws.title = "待回补批号"
    ws.append(["序号", "物品ID", "试剂名称", "规格", "品牌", "责任库",
               "材料编码", "单位", "当前库存", "最低库存", "年用量", "优先级"])
    for i, r in enumerate(rows, 1):
        if r["annual_usage"] >= 50:
            pri = "高"
        elif r["annual_usage"] >= 15:
            pri = "中"
        else:
            pri = "低"
        ws.append([i, r["id"], r["name"], r["spec"], r["brand"], r["library"],
                   r["material_code"], r["unit"], r["qty"], r["min_stock"],
                   r["annual_usage"], pri])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([6, 8, 42, 24, 14, 10, 16, 8, 10, 10, 10, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "C2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row[11].value == "高":
            for cell in row:
                cell.fill = TOP_FILL

    ws2 = wb.create_sheet("汇总")
    ws2.append(["指标", "值"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    total_items = len(set(r["item_id"] for r in stock))
    with_batch = len({r["item_id"] for r in stock if (r["batch_no"] or "").strip()})
    ws2.append(["库存物品总数", total_items])
    ws2.append(["已有批号", with_batch])
    ws2.append(["无批号", total_items - with_batch])
    ws2.append([f"其中「试剂」无批号", len(rows) if not all_types else "(未筛选)"])
    ws2.append(["高优先级(年用量≥50)", sum(1 for r in rows if r["annual_usage"] >= 50)])
    ws2.append(["中优先级(15~49)", sum(1 for r in rows if 15 <= r["annual_usage"] < 50)])
    ws2.append(["低优先级(<15)", sum(1 for r in rows if r["annual_usage"] < 15)])
    ws2.append(["说明", "批号/效期只在「到货接收」录入，盘库不填。"
                        "本清单按年用量降序，下次到货时优先补录高优先级项目。"])
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 70

    wb.save(OUT)
    print("已生成:", OUT)
    print(f"无批号试剂 {len(rows)} 个（共 {total_items} 个物品，{with_batch} 个已有批号）")


if __name__ == "__main__":
    main()
