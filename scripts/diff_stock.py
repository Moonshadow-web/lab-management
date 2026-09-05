"""对比库存快照与当前库存，生成整理前后对照表。

用法：python scripts/diff_stock.py <备份快照.json>
"""
import glob
import json
import os
import sys
import urllib.parse
import urllib.request
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

H = "http://lab-management-282724-9-1408547492.sh.run.tcloudbase.com"
OUT = "outputs/实时库存整理前后对照.xlsx"
HEAD_FILL = PatternFill("solid", fgColor="1a365d")
CHG_FILL = PatternFill("solid", fgColor="FFF2CC")


def login():
    data = urllib.parse.urlencode(
        {"username": "jinzizheng", "password": "Jzz6827556"}).encode()
    req = urllib.request.Request(H + "/api/v1/auth/login", data=data)
    return json.load(urllib.request.urlopen(req, timeout=60))["access_token"]


def get(tok, path):
    req = urllib.request.Request(H + path,
                                 headers={"Authorization": "Bearer " + tok})
    return json.load(urllib.request.urlopen(req, timeout=300))


def paged(tok, path, size=200):
    out, page = [], 1
    while True:
        d = get(tok, f"{path}&page={page}&page_size={size}")
        out += d["items"]
        if len(out) >= d["total"] or not d["items"]:
            break
        page += 1
    return out


def main():
    snap_path = sys.argv[1] if len(sys.argv) > 1 else sorted(
        glob.glob("outputs/stock_backup_*.json"))[-1]
    snap = json.load(open(snap_path, encoding="utf-8"))
    tok = login()
    names = {i["id"]: i for i in paged(tok, "/api/v1/reagent/items?page=1")}
    now = paged(tok, "/api/v1/reagent/stock?page=1")

    def agg(rows):
        m = defaultdict(lambda: {"qty": 0, "rows": 0, "batches": []})
        for r in rows:
            k = r["item_id"]
            m[k]["qty"] += int(r["quantity"] or 0)
            m[k]["rows"] += 1
            m[k]["batches"].append(f"{r['batch_no'] or '(空)'}×{r['quantity']}")
        return m

    before, after = agg(snap["stock"]), agg(now)
    all_ids = sorted(set(before) | set(after))

    wb = Workbook()
    ws = wb.active
    ws.title = "整理前后对照"
    ws.append(["物品ID", "试剂 / 耗材名称", "规格", "库",
               "整理前行数", "整理前数量", "整理前明细",
               "整理后行数", "整理后数量", "整理后明细", "变化"])
    for iid in all_ids:
        b, a = before.get(iid), after.get(iid)
        bq = b["qty"] if b else 0
        aq = a["qty"] if a else 0
        if bq == aq and (b["rows"] if b else 0) == (a["rows"] if a else 0):
            continue
        it = names.get(iid, {})
        delta = aq - bq
        ws.append([
            iid, it.get("name", f"(id={iid})"), it.get("spec", ""),
            it.get("library", ""),
            b["rows"] if b else 0, bq, "；".join(b["batches"]) if b else "",
            a["rows"] if a else 0, aq, "；".join(a["batches"]) if a else "",
            f"{delta:+d}" if delta else "行数变化",
        ])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate([8, 40, 22, 10, 10, 10, 46, 10, 10, 30, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    changed = ws.max_row - 1
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        row[10].fill = CHG_FILL

    ws2 = wb.create_sheet("汇总")
    ws2.append(["指标", "整理前", "整理后"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    ws2.append(["库存总行数", len(snap["stock"]), len(now)])
    ws2.append(["物品数", len(before), len(after)])
    ws2.append(["有变化的物品数", "—", changed])
    ws2.append(["存在多行的物品数",
                sum(1 for k in before if before[k]["rows"] > 1),
                sum(1 for k in after if after[k]["rows"] > 1)])
    ws2.append(["收货单张数", len(snap["receivings"]), "19（删除重复单 RCV-2026-08-660）"])
    ws2.append(["快照文件", os.path.basename(snap_path), snap.get("backup_at", "")])
    for i, w in enumerate([26, 16, 46], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUT)
    print("已生成:", OUT)
    print(f"有变化的物品 {changed} 个；库存行 {len(snap['stock'])} -> {len(now)}")


if __name__ == "__main__":
    main()
