"""导出 reagent_stock 脏数据整理预演报告（Excel），供人工确认后再执行。

用法：
  python scripts/report_stock_normalize.py            # 默认 dry_run，输出 xlsx
  python scripts/report_stock_normalize.py --apply    # 确认后实际执行整理
"""
import json
import sys
import urllib.parse
import urllib.request

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

H = "http://lab-management-282724-9-1408547492.sh.run.tcloudbase.com"
LOGIN_URL = H + "/api/v1/auth/login"
USERNAME = "jinzizheng"
PASSWORD = "Jzz6827556"
OUT = "outputs/实时库存整理预演.xlsx"


def login():
    data = urllib.parse.urlencode({"username": USERNAME, "password": PASSWORD}).encode()
    req = urllib.request.Request(LOGIN_URL, data=data)
    return json.load(urllib.request.urlopen(req, timeout=60))["access_token"]


def call(tok, path, method="GET"):
    url = H + path
    req = urllib.request.Request(url, method=method,
                                 headers={"Authorization": "Bearer " + tok})
    return json.load(urllib.request.urlopen(req, timeout=300))


def main():
    apply_changes = "--apply" in sys.argv
    tok = login()

    # 试剂名称映射
    names = {}
    page = 1
    while True:
        d = call(tok, f"/api/v1/reagent/items?page={page}&page_size=200")
        for it in d["items"]:
            names[it["id"]] = it
        if page * 200 >= d["total"]:
            break
        page += 1

    res = call(tok, "/api/v1/reagent/stock/_normalize?dry_run=" +
               ("false" if apply_changes else "true"), method="POST")

    wb = Workbook()
    ws = wb.active
    ws.title = "整理预演"
    head = (["序号", "类型", "物品ID", "试剂 / 耗材名称", "规格", "库",
             "原批号", "目标批号", "原有行数", "原数量合计", "整理后数量", "说明"])
    ws.append(head)
    fill = PatternFill("solid", fgColor="1a365d")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 现有库存，用于算原数量合计
    stock_rows = []
    page = 1
    while True:
        d = call(tok, f"/api/v1/reagent/stock?page={page}&page_size=200")
        stock_rows += d["items"]
        if len(stock_rows) >= d["total"]:
            break
        page += 1
    from collections import defaultdict
    by_item = defaultdict(list)
    for r in stock_rows:
        by_item[r["item_id"]].append(r)

    for i, c in enumerate(res["changes"], 1):
        iid = c["item_id"]
        it = names.get(iid, {})
        rows = by_item.get(iid, [])
        total_qty = sum(int(r["quantity"] or 0) for r in rows)
        if c["action"] == "dedupe":
            ws.append([i, "重复行合并", iid, it.get("name", f"(id={iid})"),
                       it.get("spec", ""), it.get("library", ""),
                       c["batch_no"] or "(空批号)", c["batch_no"] or "(空批号)",
                       c["rows"], total_qty, c["quantity"],
                       "同一试剂在盘库单里跨分组被重复提交，多行值相同，合并为1行（取最大值）"])
        else:
            ws.append([i, "空批号并入默认批号", iid, it.get("name", f"(id={iid})"),
                       it.get("spec", ""), it.get("library", ""),
                       "(空批号)", c["to"], c.get("rows", 2), total_qty,
                       c.get("result") or c.get("quantity"),
                       "盘库未填批号留下的空批号行，回填该试剂默认批号后并入（数量相加）"])

    widths = [6, 18, 9, 40, 24, 10, 16, 16, 10, 12, 12, 56]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("汇总")
    ws2.append(["项目", "值"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    ws2.append(["整理前行数", res["rows_before"]])
    ws2.append(["整理后行数", res["rows_after"]])
    ws2.append(["变更条目数", res["count"]])
    ws2.append(["模式", "已执行" if not res["dry_run"] else "仅预演（未改数据）"])
    if not res["dry_run"]:
        ws2.append(["删除行数", res.get("deleted_rows", 0)])
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 30

    wb.save(OUT)
    print("已生成:", OUT)
    print("整理前", res["rows_before"], "行 →", res["rows_after"], "行，变更", res["count"], "条")
    print("模式:", "已执行" if not res["dry_run"] else "仅预演")


if __name__ == "__main__":
    main()
