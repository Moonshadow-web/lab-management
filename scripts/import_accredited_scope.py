"""解析《生免组申请认可的能力范围.xlsx》并批量导入线上 accredited_scopes 表。

列映射（1-based，已对照表头第 1/2 行核实）：
  A=1  序号(seq)
  B=2  检验（检查）项目(item_name)
  C=3  样品类型(sample_type)
  D=4  检验（检查）方法(method_name，旧值，留待系统中关联)
  E=5  设备(instrument_name，旧值)
  F=6  试剂(reagent_name，旧值)
  G=7  校准品(calibrator)
  H=8  说明(description)
  I=9  备注(remark)
  J=10 正确度(perf_correctness)
  K=11 精密度(perf_precision)
  L=12 线性(perf_linearity)
  O=15 可报告范围(perf_reportable)
  Q=17 其他(perf_other)  —— 表头把"5.其他"标在 R(18)，但数据实际落在 Q(17)，R 全空

分组：A 列非空、B 列为空 → 分组行。
  - 首字母 1 个（如 "A 检验医学"）→ 一级分类 category_l1，并清空二级
  - 首字母 >=2 个（如 "AA 临床血液学"）→ 二级分类 category_l2

方法/试剂/仪器先按 Excel 旧文本存入 *_name，*_id 留空；
后续在「认可能力范围」前端用下拉关联到系统中正式实体。

用法：
  python import_accredited_scope.py            # 仅解析预览（dry-run）
  python import_accredited_scope.py --apply    # 真正调用线上 API 批量导入（replace=True）
"""
from __future__ import annotations

import argparse
import json
import re
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl

BASE = "http://lab-management-282724-9-1408547492.sh.run.tcloudbase.com/api/v1"
USERNAME = "jinzizheng"
PASSWORD = "Jzz6827556"

DEFAULT_XLSX = r"D:/民航总医院/15189/生免认可申请附表/生免组申请认可的能力范围.xlsx"

# 列（1-based）
COL = {
    "seq": 1, "item_name": 2, "sample_type": 3, "method_name": 4,
    "instrument_name": 5, "reagent_name": 6, "calibrator": 7,
    "description": 8, "remark": 9,
    "perf_correctness": 10, "perf_precision": 11, "perf_linearity": 12,
    "perf_reportable": 15, "perf_other": 17,
}


def _cell(ws, row: int, col: int):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, (int, float)):
        return str(v)
    return str(v).strip()


def parse_xlsx(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[dict] = []
    cur_l1 = ""
    cur_l2 = ""
    for r in range(3, ws.max_row + 1):
        a = _cell(ws, r, COL["seq"])
        b = _cell(ws, r, COL["item_name"])
        # 分组行：A 非空、B 空
        if a and not b:
            m = re.match(r"^([A-Z]+)\s", a)
            n = len(m.group(1)) if m else 0
            if n <= 1:
                cur_l1 = a
                cur_l2 = ""
            else:
                cur_l2 = a
            continue
        if not b:
            continue  # 空行跳过
        item = {
            "category_l1": cur_l1,
            "category_l2": cur_l2,
            "seq": a,
            "item_name": b,
            "sample_type": _cell(ws, r, COL["sample_type"]),
            "method_name": _cell(ws, r, COL["method_name"]),
            "instrument_name": _cell(ws, r, COL["instrument_name"]),
            "reagent_name": _cell(ws, r, COL["reagent_name"]),
            "calibrator": _cell(ws, r, COL["calibrator"]),
            "description": _cell(ws, r, COL["description"]),
            "remark": _cell(ws, r, COL["remark"]),
            "perf_correctness": _cell(ws, r, COL["perf_correctness"]),
            "perf_precision": _cell(ws, r, COL["perf_precision"]),
            "perf_linearity": _cell(ws, r, COL["perf_linearity"]),
            "perf_reportable": _cell(ws, r, COL["perf_reportable"]),
            "perf_other": _cell(ws, r, COL["perf_other"]),
        }
        rows.append(item)
    return rows


def api_request(method: str, path: str, data=None, token: str = "") -> dict | list:
    url = BASE + path
    headers = {"Accept": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    body = None
    if data is not None:
        if method == "POST" and path == "/auth/login":
            body = urllib.parse.urlencode(data).encode("utf-8")
            headers["Content-Type"] = "application/x-www-form-urlencoded"
        else:
            body = json.dumps(data, ensure_ascii=False).encode("utf-8")
            headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        text = e.read().decode("utf-8", "ignore")
        raise RuntimeError(f"{method} {path} {e.code}: {text[:400]}")


def login() -> str:
    r = api_request("POST", "/auth/login", {"username": USERNAME, "password": PASSWORD})
    return r["access_token"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=DEFAULT_XLSX)
    ap.add_argument("--apply", action="store_true", help="真正导入（默认仅预览）")
    args = ap.parse_args()

    items = parse_xlsx(args.xlsx)
    print(f"解析得到 {len(items)} 条认可能力范围项目")
    # 预览前 5 条
    for it in items[:5]:
        print(" -", it["category_l1"], "/", it["category_l2"], "|",
              it["seq"], it["item_name"], "| 方法:", it["method_name"],
              "| 设备:", it["instrument_name"][:30])
    if not args.apply:
        print("\n[dry-run] 未导入。加 --apply 执行线上批量导入。")
        return

    token = login()
    resp = api_request("POST", "/accredited-scope/batch?replace=true",
                       data=items, token=token)
    print("导入结果:", resp)


if __name__ == "__main__":
    main()
