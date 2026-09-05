"""试剂库存溯源 + 重复收货单嫌疑扫描。

产出 outputs/试剂库存溯源与重复单据嫌疑.xlsx，包含：
  1) 指定试剂的库存来源时间线（到货接收 + 盘库 → 当前库存行）
  2) 全部收货单一览（含确认时间/确认人/创建人）
  3) 重复收货单嫌疑（同收货日期 + 同送货人 + 同物品 + 同数量）
"""
import json
import sys
import urllib.parse
import urllib.request
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

H = "http://lab-management-282724-9-1408547492.sh.run.tcloudbase.com"
OUT = "outputs/试剂库存溯源与重复单据嫌疑.xlsx"
HEAD_FILL = PatternFill("solid", fgColor="1a365d")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")


def login():
    data = urllib.parse.urlencode(
        {"username": "jinzizheng", "password": "Jzz6827556"}).encode()
    req = urllib.request.Request(H + "/api/v1/auth/login", data=data)
    return json.load(urllib.request.urlopen(req, timeout=60))["access_token"]


def get(tok, path):
    req = urllib.request.Request(H + path,
                                 headers={"Authorization": "Bearer " + tok})
    return json.load(urllib.request.urlopen(req, timeout=300))


def paged(tok, path, size=100):
    out, page = [], 1
    while True:
        d = get(tok, f"{path}&page={page}&page_size={size}")
        out += d["items"]
        if len(out) >= d["total"] or not d["items"]:
            break
        page += 1
    return out


def style_header(ws, widths):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def main():
    targets = [int(x) for x in sys.argv[1:]] or [74, 112, 83, 315]
    tok = login()
    names = {i["id"]: i for i in paged(tok, "/api/v1/reagent/items?page=1", 200)}
    stock = paged(tok, "/api/v1/reagent/stock?page=1", 200)
    recs = paged(tok, "/api/v1/reagent/receivings?page=1")
    checks = paged(tok, "/api/v1/reagent/inventory-checks?page=1")

    by_item = defaultdict(list)
    for s in stock:
        by_item[s["item_id"]].append(s)

    wb = Workbook()

    # ── Sheet1：目标试剂溯源时间线 ──
    ws = wb.active
    ws.title = "库存溯源"
    ws.append(["试剂", "规格", "日期", "环节", "单据号", "批号", "效期",
               "入库/余量", "该批号当时余额", "备注"])
    for iid in targets:
        it = names.get(iid, {})
        nm = it.get("name", f"(id={iid})")
        spec = it.get("spec", "")
        events = []
        for r in recs:
            for li in (r.get("items") or []):
                if li["item_id"] == iid:
                    events.append((r["receipt_date"], 0, r, li, "到货接收"))
        for c in checks:
            for li in (c.get("items") or []):
                if li["item_id"] == iid:
                    events.append((c["check_date"], 1, c, li, "盘库"))
        events.sort(key=lambda x: (x[0], x[1], x[2].get("id", 0)))

        # 模拟余额
        bal = defaultdict(int)
        for _dt, _o, doc, li, kind in events:
            b = li["batch_no"] or "(空)"
            if kind == "到货接收":
                if doc.get("is_confirmed"):
                    bal[b] += int(li["quantity"] or 0)
                    note = f"已确认（{doc.get('confirmed_at', '')[:10]}，{doc.get('confirmed_by', '')}）"
                else:
                    note = "未确认 → 未入库存"
            else:
                note = f"盘库覆盖（{doc.get('check_type')}）"
                if b == "(空)":
                    for k in list(bal):
                        bal[k] = 0
                bal[b] = int(li["recorded_quantity"] or 0)
            qty = int(li["quantity"] if kind == "到货接收" else li["recorded_quantity"])
            ws.append([nm, spec, str(_dt), kind,
                       doc.get("receipt_no") or f"盘库#{doc.get('id')}",
                       li["batch_no"] or "(空)", li["expiry_date"] or "",
                       qty, bal[b], note])
        for s in by_item.get(iid, []):
            ws.append([nm, spec, "", "当前库存行", f"stock id={s['id']}",
                       s["batch_no"] or "(空)", s["expiry_date"] or "",
                       s["quantity"], s["quantity"], "前端显示的余额"])
        ws.append(["", "", "", "", "", "", "", "", "", ""])
    style_header(ws, [34, 20, 12, 12, 20, 14, 12, 10, 14, 40])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Sheet2：全部收货单 ──
    ws2 = wb.create_sheet("收货单一览")
    ws2.append(["收货单号", "id", "收货日期", "送货人", "创建人", "是否已确认",
                "确认时间", "确认人", "细项数", "物品清单"])
    for r in sorted(recs, key=lambda x: (x["receipt_date"], x["id"]), reverse=True):
        items_str = "；".join(
            f"{names.get(li['item_id'], {}).get('name', li['item_id'])[:16]}"
            f"({li['batch_no']})×{li['quantity']}"
            for li in sorted(r.get("items") or [], key=lambda x: x["item_id"]))
        ws2.append([r["receipt_no"], r["id"], r["receipt_date"],
                    r.get("delivery_person", ""), r.get("created_by", ""),
                    "是" if r.get("is_confirmed") else "否",
                    (r.get("confirmed_at") or "")[:19], r.get("confirmed_by", ""),
                    len(r.get("items") or []), items_str])
    style_header(ws2, [20, 6, 12, 12, 14, 10, 20, 12, 8, 80])
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        if row[5].value == "否":
            for cell in row:
                cell.fill = WARN_FILL

    # ── Sheet3：重复收货单嫌疑 ──
    ws3 = wb.create_sheet("重复单据嫌疑")
    ws3.append(["收货日期", "送货人", "物品", "数量", "涉及单据", "批号差异", "判断"])
    sig = defaultdict(list)  # (date, person, item_id, qty) -> [(单号, 批号)]
    for r in recs:
        for li in (r.get("items") or []):
            key = (r["receipt_date"], r.get("delivery_person", ""),
                   li["item_id"], int(li["quantity"] or 0))
            sig[key].append((r["receipt_no"], r.get("is_confirmed"),
                             li["batch_no"] or "", r.get("created_by", "")))
    n = 0
    for key, lst in sorted(sig.items()):
        if len(lst) < 2:
            continue
        dt, person, iid, qty = key
        batches = [x[2] for x in lst]
        uniq = sorted(set(batches))
        verdict = ("批号完全一致 → 高度疑似重复录入"
                   if len(uniq) == 1 else
                   f"批号不同（{' / '.join(uniq)}）→ 疑似重复录入且有一处批号笔误")
        n += 1
        ws3.append([dt, person, names.get(iid, {}).get("name", f"(id={iid})"),
                    qty,
                    "；".join(f"{x[0]}({'已确认' if x[1] else '未确认'},建单人{x[3]})"
                              for x in lst),
                    " / ".join(b or "(空)" for b in batches), verdict])
    if n == 0:
        ws3.append(["（未发现同日+同送货人+同物品+同数量的重复单据）", "", "", "", "", "", ""])
    style_header(ws3, [12, 12, 34, 8, 52, 26, 40])
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        for cell in row:
            cell.fill = WARN_FILL

    wb.save(OUT)
    print("已生成:", OUT)
    print("重复单据嫌疑:", n, "组")


if __name__ == "__main__":
    main()
